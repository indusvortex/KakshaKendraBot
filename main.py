import os
import csv
import io
import json
import base64
import secrets
from contextlib import asynccontextmanager
from datetime import datetime, timezone, timedelta
from pathlib import Path
from fastapi import FastAPI, Request, HTTPException, Response, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from dotenv import load_dotenv

import database
from utils import (
    generate_ai_response,
    send_whatsapp_message,
    send_whatsapp_template,
    upload_media_to_whatsapp,
    send_whatsapp_media,
    ai_stats,
    _get_groq_clients,
)

# India Standard Time = UTC+5:30
IST = timezone(timedelta(hours=5, minutes=30))

# Server start timestamp — used for uptime calculation in admin settings page
SERVER_STARTED_AT = datetime.now(timezone.utc)


# ============================================================
# Web Push (VAPID) — for background notifications via service worker
# ============================================================
VAPID_KEYS_FILE = Path(database.DB_PATH).parent / "vapid_keys.json"


def _get_or_create_vapid_keys() -> tuple[str, str]:
    """
    Returns (public_key_b64url, private_key_pem).
    Order of preference:
      1) VAPID_PUBLIC_KEY + VAPID_PRIVATE_KEY env vars
      2) Cached keys on disk (persistent volume)
      3) Generate a new pair and persist to disk
    """
    pub_env = os.getenv("VAPID_PUBLIC_KEY")
    priv_env = os.getenv("VAPID_PRIVATE_KEY")
    if pub_env and priv_env:
        return pub_env.strip(), priv_env.strip().replace("\\n", "\n")

    if VAPID_KEYS_FILE.exists():
        try:
            data = json.loads(VAPID_KEYS_FILE.read_text())
            return data["public"], data["private"]
        except Exception:
            pass

    # Generate fresh ECDSA P-256 key pair
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.primitives import serialization

    priv_key = ec.generate_private_key(ec.SECP256R1())
    private_pem = priv_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption(),
    ).decode()

    public_numbers = priv_key.public_key().public_numbers()
    raw_public = b"\x04" + public_numbers.x.to_bytes(32, "big") + public_numbers.y.to_bytes(32, "big")
    public_b64 = base64.urlsafe_b64encode(raw_public).decode().rstrip("=")

    try:
        VAPID_KEYS_FILE.write_text(json.dumps({"public": public_b64, "private": private_pem}))
        print(f"[VAPID] Generated new keys, saved to {VAPID_KEYS_FILE}")
    except Exception as e:
        print(f"[VAPID] Could not persist keys to disk: {e}")
    return public_b64, private_pem


VAPID_PUBLIC, VAPID_PRIVATE = _get_or_create_vapid_keys()
VAPID_CLAIMS_SUB = os.getenv("VAPID_SUBJECT", "mailto:admin@kakshakendra.com")


def send_web_push(title: str, body: str, url: str = "/admin", tag: str = "kk-default", role: str | None = None):
    """
    Sends a Web Push to subscribed browsers.
    role=None -> all subscribers
    role='team' -> only team browsers (for new lead alerts)
    role='super_admin' -> only admin browsers (for escalations)
    """
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        print("[Push] pywebpush not installed; skipping web push.")
        return

    payload = json.dumps({
        "title": title,
        "body": body,
        "url": url,
        "tag": tag,
        "icon": "/favicon.ico",
    })

    subs = database.get_push_subscriptions(role)
    if not subs:
        print(f"[Push] No subscriptions for role={role!r}")
        return

    sent, failed = 0, 0
    for sub in subs:
        try:
            webpush(
                subscription_info={
                    "endpoint": sub["endpoint"],
                    "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]},
                },
                data=payload,
                vapid_private_key=VAPID_PRIVATE,
                vapid_claims={"sub": VAPID_CLAIMS_SUB},
            )
            sent += 1
        except WebPushException as e:
            failed += 1
            # 410 = subscription expired, remove it
            if e.response is not None and e.response.status_code in (404, 410):
                database.remove_push_subscription(sub["endpoint"])
                print(f"[Push] Removed expired subscription")
            else:
                print(f"[Push] Failed: {e}")
        except Exception as e:
            failed += 1
            print(f"[Push] Unexpected error: {e}")

    print(f"[Push] Sent: {sent}, Failed: {failed}")


def to_ist(utc_str: str | None, fmt: str = "%d %b %Y, %I:%M %p") -> str:
    """Converts SQLite-stored UTC timestamp string to IST formatted string."""
    if not utc_str:
        return ""
    try:
        dt_utc = datetime.strptime(utc_str.split(".")[0], "%Y-%m-%d %H:%M:%S")
        dt_utc = dt_utc.replace(tzinfo=timezone.utc)
        return dt_utc.astimezone(IST).strftime(fmt)
    except Exception:
        return utc_str


_CLASS_PATTERNS = [
    # "Class 10", "class10", "Class - 10"
    r'class[\s\-]*(\d{1,2})',
    # "10th class", "10 class"
    r'(\d{1,2})(?:st|nd|rd|th)?\s*class',
    # "10th", "9th" standalone
    r'\b(\d{1,2})(?:st|nd|rd|th)\b',
    # Hindi "10वीं", "9वी"
    r'(\d{1,2})\s*(?:वीं|वी|th\s|tha)',
    # Pre-school
    r'(pre[\s\-]*primary|primary|nursery|ukg|lkg)',
    # Junior / Secondary keywords
    r'(junior|secondary|sr\.?\s*secondary|senior\s*secondary|foundation)',
]


def _detect_class(text: str) -> str:
    """Extracts a class label from a student's message. Returns '' if not found."""
    if not text:
        return ""
    import re as _re
    lowered = text.lower()
    for pat in _CLASS_PATTERNS:
        m = _re.search(pat, lowered)
        if m:
            v = m.group(1)
            if v.isdigit():
                n = int(v)
                if 1 <= n <= 12:
                    return f"Class {n}"
            else:
                return v.title()
    return ""


def _detect_source(text: str) -> str:
    """Guesses how the student found us based on their first message."""
    if not text:
        return "WhatsApp"
    lowered = text.lower()
    if any(w in lowered for w in ["ad", "advertisement", "facebook", "instagram", "fb", "insta", "saw your"]):
        return "Meta Ad"
    if any(w in lowered for w in ["referral", "referred", "friend", "recommend", "told"]):
        return "Referral"
    if any(w in lowered for w in ["website", "google", "search"]):
        return "Website / Google"
    if any(w in lowered for w in ["story", "reel", "post"]):
        return "Instagram"
    return "WhatsApp"


def _sync_to_google_sheets(
    sender_id: str,
    name: str | None,
    message: str,
    role: str,
    is_new_lead: bool = False,
):
    """
    Sends CRM-style data to a Google Apps Script web app.
    Sheet columns: Naam | Class | Phone | Source | Status | Next Call | Notes
    Apps Script preserves Source/Status/Next Call/Notes once admin sets them manually.
    """
    webhook_url = os.getenv("GOOGLE_SHEETS_WEBHOOK", "").strip()
    if not webhook_url:
        return

    payload = {
        "phone": "+" + sender_id,
        "naam": (name or "Unknown").strip(),
        "class": _detect_class(message),
        "source": _detect_source(message) if is_new_lead else "",
        "status": "New" if is_new_lead else "",
        "last_message": message[:200],
        "timestamp": datetime.now(IST).strftime("%Y-%m-%d %H:%M"),
        "is_new_lead": is_new_lead,
    }

    try:
        import requests
        requests.post(webhook_url, json=payload, timeout=4)
        print(f"[Sheets] Synced +{sender_id} (new={is_new_lead})")
    except Exception as e:
        print(f"[Sheets] Sync failed for +{sender_id}: {e}")


def _whatsapp_broadcast(numbers_csv: str, message_text: str, label: str = "Notify"):
    """
    Sends a WhatsApp text to a comma-separated list of numbers (no '+' needed).
    Silently skips if list is empty.
    24h-window note: WhatsApp only allows free-form messages to numbers
    that have messaged the bot in the last 24h, OR via an approved template.
    """
    if not numbers_csv.strip():
        return
    numbers = [n.strip().lstrip("+") for n in numbers_csv.split(",") if n.strip()]
    for num in numbers:
        try:
            result = send_whatsapp_message(num, message_text)
            if result is not None:
                print(f"[{label}] Sent WhatsApp -> +{num}")
            else:
                print(f"[{label}] Failed -> +{num} (24h window or token issue)")
        except Exception as e:
            print(f"[{label}] Exception -> +{num}: {e}")


def _notify_team_of_new_lead(sender_id: str, name: str | None, first_message: str,
                             class_label: str = "", source: str = ""):
    """
    Sends a WhatsApp 'new student lead' alert to the TEAM numbers.
    Configure TEAM_NOTIFY_NUMBERS env var (comma-separated, with country code, no +).
    """
    team_numbers = os.getenv("TEAM_NOTIFY_NUMBERS", "").strip()
    if not team_numbers:
        return

    name_display = (name or "Unknown").strip()
    msg_preview = first_message.strip()[:200]
    notification = (
        f"🔔 *NEW LEAD — Please call!*\n"
        f"\n"
        f"👤 *Name:* {name_display}\n"
        f"📱 *Number:* +{sender_id}\n"
        f"🎓 *Class:* {class_label or 'TBD'}\n"
        f"📍 *Source:* {source or 'WhatsApp'}\n"
        f"💬 *First message:* {msg_preview!r}\n"
        f"\n"
        f"⚡ Call within 2 min or admin will be alerted.\n"
        f"📊 Open dashboard to mark as 'Called'."
    )
    _whatsapp_broadcast(team_numbers, notification, label="Team-Notify")


def _notify_admin_of_new_student(sender_id: str, name: str | None, first_message: str):
    """
    Sends a WhatsApp 'new student' summary to ADMIN numbers (CC).
    Configure ADMIN_NOTIFY_NUMBER env var (comma-separated, with country code, no +).
    Note: WhatsApp 24h rule applies.
    """
    admin_numbers = os.getenv("ADMIN_NOTIFY_NUMBER", "").strip()
    if not admin_numbers:
        return

    name_display = (name or "Unknown").strip()
    msg_preview = first_message.strip()[:200]
    notification = (
        f"🔔 *New Student Alert!*\n"
        f"\n"
        f"A new student just messaged the bot:\n"
        f"\n"
        f"👤 *Name:* {name_display}\n"
        f"📱 *Number:* +{sender_id}\n"
        f"💬 *First message:* {msg_preview!r}\n"
        f"\n"
        f"📊 Team has been notified to call."
    )
    _whatsapp_broadcast(admin_numbers, notification, label="Admin-Notify")


def _notify_admin_lead_pending(sender_id: str, name: str, class_label: str, mins_waiting: int):
    """
    Sends a WhatsApp message to ADMIN numbers when a lead has been waiting
    (called every 5 minutes from the reminder loop until it's marked called).
    """
    admin_numbers = os.getenv("ADMIN_NOTIFY_NUMBER", "").strip()
    if not admin_numbers:
        return

    notification = (
        f"📞 *Lead still pending — {mins_waiting} min waiting*\n"
        f"\n"
        f"👤 *Name:* {name}\n"
        f"📱 *Number:* +{sender_id}\n"
        f"🎓 *Class:* {class_label or 'TBD'}\n"
        f"\n"
        f"⚠️ Team has not marked this lead as called yet.\n"
        f"📊 Please follow up or check with the team."
    )
    _whatsapp_broadcast(admin_numbers, notification, label="Admin-Reminder")


def _notify_team_lead_pending(sender_id: str, name: str, class_label: str, mins_waiting: int):
    """
    Sends a WhatsApp 'have you called?' reminder to TEAM numbers every 5 minutes
    until the lead is marked as called.
    """
    team_numbers = os.getenv("TEAM_NOTIFY_NUMBERS", "").strip()
    if not team_numbers:
        return

    notification = (
        f"⏰ *Reminder — Have you called this lead?*\n"
        f"\n"
        f"👤 *Name:* {name}\n"
        f"📱 *Number:* +{sender_id}\n"
        f"🎓 *Class:* {class_label or 'TBD'}\n"
        f"⏱️ *Waiting:* {mins_waiting} min\n"
        f"\n"
        f"📊 Open the dashboard and tap '✓ Mark Called' once done.\n"
        f"🚨 If you don't act, admin will be notified."
    )
    _whatsapp_broadcast(team_numbers, notification, label="Team-Reminder")

load_dotenv()

# Verify Token used by Meta to verify webhook
VERIFY_TOKEN = os.getenv("VERIFY_TOKEN", "my_secure_verify_token")

# Admin credentials
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "kakshakendra2026")
# Team credentials (limited access — chat only, no settings)
TEAM_USERNAME = os.getenv("TEAM_USERNAME", "Team")
TEAM_PASSWORD = os.getenv("TEAM_PASSWORD", "kakshakendra2026")
security = HTTPBasic()


def _check_admin(c: HTTPBasicCredentials) -> bool:
    return (
        secrets.compare_digest(c.username, ADMIN_USERNAME)
        and secrets.compare_digest(c.password, ADMIN_PASSWORD)
    )


def _check_team(c: HTTPBasicCredentials) -> bool:
    return (
        secrets.compare_digest(c.username, TEAM_USERNAME)
        and secrets.compare_digest(c.password, TEAM_PASSWORD)
    )


def verify_user(credentials: HTTPBasicCredentials = Depends(security)) -> dict:
    """
    Allows both super-admin and team to access dashboard chat features.
    Returns {'username', 'role'} where role is 'super_admin' or 'team'.
    """
    if _check_admin(credentials):
        return {"username": credentials.username, "role": "super_admin"}
    if _check_team(credentials):
        return {"username": credentials.username, "role": "team"}
    raise HTTPException(
        status_code=401,
        detail="Invalid credentials",
        headers={"WWW-Authenticate": "Basic"},
    )


def verify_admin(credentials: HTTPBasicCredentials = Depends(security)) -> dict:
    """Super-admin only — used for settings, exports, system actions."""
    if _check_admin(credentials):
        return {"username": credentials.username, "role": "super_admin"}
    raise HTTPException(
        status_code=401,
        detail="Super admin access required",
        headers={"WWW-Authenticate": "Basic"},
    )

# Track processed message IDs to avoid duplicate processing
# (WhatsApp can retry webhooks, sending the same message twice)
processed_message_ids: set = set()

async def reminder_loop():
    """
    Background loop that:
    - Re-pushes pending lead reminders to TEAM every 5 minutes.
    - Escalates to SUPER_ADMIN if a lead has been waiting > 2 minutes
      and no team member has responded yet.
    Runs forever in the background.
    """
    import asyncio
    REMINDER_INTERVAL_SECONDS = 300   # 5 minutes between team pushes
    ESCALATE_AFTER_SECONDS = 120      # 2 minutes -> notify admin
    POLL_EVERY_SECONDS = 30

    while True:
        try:
            pending = database.get_pending_lead_reminders()
            now_utc = datetime.now(timezone.utc)
            for lead in pending:
                sender_id = lead["sender_id"]
                naam = lead.get("naam") or "Unknown"
                class_label = lead.get("class_label") or ""

                # Parse timestamps stored as SQLite UTC strings
                def _parse(ts: str | None):
                    if not ts:
                        return None
                    try:
                        return datetime.strptime(
                            ts.split(".")[0], "%Y-%m-%d %H:%M:%S"
                        ).replace(tzinfo=timezone.utc)
                    except Exception:
                        return None

                last_remind = _parse(lead.get("last_reminded_at"))
                admin_notified = _parse(lead.get("admin_notified_at"))
                created = _parse(lead.get("created_at")) or now_utc

                age_since_last_remind = (now_utc - (last_remind or created)).total_seconds()
                age_since_creation = (now_utc - created).total_seconds()

                # 1. Every 5 minutes: push + WhatsApp ping to TEAM, plus WhatsApp CC to ADMIN
                if last_remind is None or age_since_last_remind >= REMINDER_INTERVAL_SECONDS:
                    mins_waiting = int(age_since_creation / 60)

                    # Web push to team browsers (real-time)
                    send_web_push(
                        title="📞 Reminder — call this lead!",
                        body=f"{naam} ({class_label or 'class TBD'}) — still pending",
                        url=f"/admin?chat={sender_id}",
                        tag=f"remind-{sender_id}",
                        role="team",
                    )
                    # WhatsApp reminder to TEAM ("have you called yet?")
                    _notify_team_lead_pending(sender_id, naam, class_label, mins_waiting)
                    # WhatsApp CC to ADMIN ("team hasn't acted")
                    _notify_admin_lead_pending(sender_id, naam, class_label, mins_waiting)

                    database.mark_lead_reminded(sender_id)

                # 2. Escalate to admin if not handled in 2 minutes
                if admin_notified is None and age_since_creation >= ESCALATE_AFTER_SECONDS:
                    send_web_push(
                        title="🚨 ESCALATION — Team didn't call!",
                        body=f"{naam} ({class_label or 'class TBD'}) — waiting >2 min",
                        url=f"/admin?chat={sender_id}",
                        tag=f"escalate-{sender_id}",
                        role="super_admin",
                    )
                    # WhatsApp escalation to admin
                    mins_waiting = int(age_since_creation / 60)
                    _notify_admin_lead_pending(sender_id, naam, class_label, mins_waiting)
                    database.mark_lead_admin_notified(sender_id)
        except Exception as e:
            print(f"[ReminderLoop] error: {e}")
        await asyncio.sleep(POLL_EVERY_SECONDS)


@asynccontextmanager
async def lifespan(app: FastAPI):
    import asyncio
    print("Initializing Database...")
    database.init_db()
    # Start the reminder background loop
    task = asyncio.create_task(reminder_loop())
    print("[ReminderLoop] started")
    try:
        yield
    finally:
        task.cancel()

app = FastAPI(title="WhatsApp AI Coach Bot", lifespan=lifespan)


@app.get("/webhook")
async def verify_webhook(request: Request):
    """
    Meta Challenge Verification.
    WhatsApp sends a GET request here when you configure the Webhook.
    """
    mode = request.query_params.get("hub.mode")
    token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")

    if mode and token:
        if mode == "subscribe" and token == VERIFY_TOKEN:
            print("WEBHOOK_VERIFIED")
            return Response(content=challenge, media_type="text/plain")
        else:
            raise HTTPException(status_code=403, detail="Verification token mismatch")
    return Response(content="Hello World", media_type="text/plain")


@app.post("/webhook")
async def handle_whatsapp_message(request: Request):
    """
    Receives incoming messages from WhatsApp.
    """
    try:
        body = await request.json()

        if body.get("object") == "whatsapp_business_account":
            for entry in body.get("entry", []):
                for change in entry.get("changes", []):
                    value = change.get("value", {})

                    # Capture contact display name (sent in webhook payload alongside messages)
                    contacts_by_wa_id = {}
                    for contact in value.get("contacts", []):
                        wa_id = contact.get("wa_id")
                        name = contact.get("profile", {}).get("name")
                        if wa_id and name:
                            contacts_by_wa_id[wa_id] = name

                    # ----- Status events: delivery / read / failed for messages WE sent -----
                    if "statuses" in value:
                        for st in value["statuses"]:
                            st_type = st.get("status", "?")
                            msg_id = st.get("id", "")
                            recipient = st.get("recipient_id", "")
                            if st_type == "failed":
                                errors = st.get("errors", []) or []
                                for err in errors:
                                    print(
                                        f"[Status] ✗ FAILED to +{recipient} | "
                                        f"code={err.get('code')} title={err.get('title')!r} "
                                        f"message={err.get('message')!r} "
                                        f"details={err.get('error_data', {}).get('details')!r} "
                                        f"msg_id={msg_id}"
                                    )
                            else:
                                # 'sent' / 'delivered' / 'read'
                                print(f"[Status] {st_type} | +{recipient} | msg_id={msg_id}")

                    if "messages" in value:
                        for message in value["messages"]:
                            # Deduplicate: skip if we already processed this message
                            message_id = message.get("id")
                            if message_id and message_id in processed_message_ids:
                                print(f"Skipping duplicate message: {message_id}")
                                continue
                            if message_id:
                                processed_message_ids.add(message_id)
                                # Prevent unbounded growth in memory
                                if len(processed_message_ids) > 10000:
                                    processed_message_ids.clear()

                            sender_id = message["from"]

                            # Save / update the student's display name for the dashboard
                            if sender_id in contacts_by_wa_id:
                                database.upsert_contact(sender_id, contacts_by_wa_id[sender_id])

                            if message["type"] == "text":
                                message_text = message["text"]["body"]
                            elif message["type"] == "interactive":
                                interactive = message["interactive"]
                                if interactive["type"] == "button_reply":
                                    message_text = interactive["button_reply"]["title"]
                                elif interactive["type"] == "list_reply":
                                    message_text = interactive["list_reply"]["title"]
                                else:
                                    continue
                            else:
                                continue

                            print(f"Received message from {sender_id}: {message_text}")

                            # ----- Special intercept: "Call Us" button tap -----
                            # WhatsApp's reply buttons can't open the dialer directly,
                            # so when the student clicks "Call Us", we send the
                            # pre-approved Meta template (which has a real PHONE_NUMBER button).
                            if message_text.strip().lower() in {"call us", "call now", "call"}:
                                template_name = os.getenv("CALL_TEMPLATE_NAME", "kaksha_call_us")
                                template_lang = os.getenv("CALL_TEMPLATE_LANG", "en")
                                ok = send_whatsapp_template(sender_id, template_name, template_lang)
                                if ok:
                                    # Save messages only on success — let AI flow handle it on failure
                                    database.save_message(sender_id, "user", message_text)
                                    database.save_message(
                                        sender_id, "assistant",
                                        f"[Sent template: {template_name}]"
                                    )
                                    continue  # skip the AI step
                                print(f"[Template] Falling back to AI for 'Call Us' (template failed)")

                            # 1. Fetch past history BEFORE saving current message
                            #    so the current message doesn't appear twice in the AI prompt
                            history = database.get_recent_messages(sender_id, limit=10)

                            contact_name = contacts_by_wa_id.get(sender_id)

                            # If this is the very first message from this student
                            if not history:
                                detected_class = _detect_class(message_text)
                                detected_source = _detect_source(message_text)

                                # Create a lead reminder so the team needs to call this student
                                database.add_lead_reminder(
                                    sender_id=sender_id,
                                    naam=contact_name or "Unknown",
                                    class_label=detected_class,
                                    phone="+" + sender_id,
                                    source=detected_source,
                                )

                                # WhatsApp notification: TEAM gets the lead, ADMIN gets a CC
                                _notify_team_of_new_lead(
                                    sender_id, contact_name, message_text,
                                    class_label=detected_class, source=detected_source,
                                )
                                _notify_admin_of_new_student(sender_id, contact_name, message_text)

                                # Web push to TEAM browsers (they handle calling)
                                send_web_push(
                                    title="🔔 New Lead — Call required!",
                                    body=f"{contact_name or '+' + sender_id} ({detected_class or 'class TBD'}): {message_text[:60]}",
                                    url=f"/admin?chat={sender_id}",
                                    tag=f"new-{sender_id}",
                                    role="team",
                                )

                                # Sync new lead to Google Sheets
                                _sync_to_google_sheets(
                                    sender_id, contact_name, message_text,
                                    role="user", is_new_lead=True,
                                )
                            else:
                                # Existing student replying — send a softer push
                                send_web_push(
                                    title="💬 " + (contact_name or "+" + sender_id),
                                    body=message_text[:120],
                                    url=f"/admin?chat={sender_id}",
                                    tag=f"chat-{sender_id}",
                                )
                                # Sync follow-up message to Google Sheets
                                _sync_to_google_sheets(
                                    sender_id, contact_name, message_text, role="user",
                                )

                            # 2. Save current user message to DB
                            database.save_message(sender_id, "user", message_text)

                            # 3. Generate AI response (history = past only, message_text = current)
                            ai_response_text = generate_ai_response(history, message_text)

                            # 4. Send back via WhatsApp API
                            send_whatsapp_message(sender_id, ai_response_text)

                            # 5. Save AI response to DB
                            database.save_message(sender_id, "assistant", ai_response_text)

        return {"status": "success"}

    except Exception as e:
        print(f"Error processing webhook: {e}")
        # Always return 200 so WhatsApp doesn't retry endlessly
        return {"status": "error"}


@app.get("/")
def health_check():
    return {"status": "Bot is running perfectly!"}


# ============================================================
# Admin Dashboard — view all student conversations
# ============================================================

def _initials_avatar(name: str | None, phone: str) -> str:
    """Returns a small HTML <span> with initials or phone digits as a colored avatar."""
    if name and name.strip():
        parts = name.strip().split()
        initials = (parts[0][0] + (parts[-1][0] if len(parts) > 1 else "")).upper()
    else:
        initials = phone[-2:] if len(phone) >= 2 else phone
    # Hash phone -> color so each student gets a consistent color
    color_palette = ["#8b5cf6", "#10b981", "#f59e0b", "#ef4444", "#3b82f6", "#ec4899", "#14b8a6"]
    color = color_palette[sum(ord(c) for c in phone) % len(color_palette)]
    return (
        f'<span style="display:inline-flex;align-items:center;justify-content:center;'
        f'width:36px;height:36px;border-radius:50%;background:{color};color:white;'
        f'font-weight:600;font-size:13px;flex-shrink:0">{initials}</span>'
    )


def _clean_message_text(text: str):
    """
    Strips [OPTIONS]...[/OPTIONS] and [CTA_URL ...] tags from a saved bot message
    so the admin dashboard shows the human-readable body + button chips separately.
    Returns (clean_text, list_of_button_labels).
    """
    import re as _re

    buttons = []

    options_match = _re.search(r'\[OPTIONS\](.*?)\[/OPTIONS\]', text, _re.DOTALL)
    if options_match:
        for line in options_match.group(1).strip().split("\n"):
            line = line.strip()
            if line:
                buttons.append(line)
        text = _re.sub(r'\[OPTIONS\].*?\[/OPTIONS\]', '', text, flags=_re.DOTALL)

    cta_match = _re.search(r'\[CTA_URL\s+display="([^"]+)"\s+url="([^"]+)"\]', text)
    if cta_match:
        buttons.append(f'🔗 {cta_match.group(1)}')
        text = _re.sub(r'\[CTA_URL[^\]]*\]', '', text)

    return text.strip(), buttons


# ----------------- Shared CSS for the messenger-style admin -----------------
_ADMIN_CSS = """
* { box-sizing: border-box; -webkit-tap-highlight-color: transparent; }
html, body { height: 100%; height: 100dvh; }
body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    color: #e6edf3;
    margin: 0;
    overflow: hidden;
    letter-spacing: -0.01em;
    /* Animated cosmic gradient background — gives glass elements something rich to blur */
    background:
        radial-gradient(ellipse at 0% 0%, rgba(82,136,193,0.18) 0%, transparent 45%),
        radial-gradient(ellipse at 100% 100%, rgba(217,70,170,0.10) 0%, transparent 45%),
        radial-gradient(ellipse at 50% 50%, rgba(16,185,129,0.05) 0%, transparent 50%),
        linear-gradient(135deg, #0a1320 0%, #0e1621 50%, #14202d 100%);
    background-attachment: fixed;
}
a { color: #5288c1; text-decoration: none; }

/* Smooth global scrollbar */
*::-webkit-scrollbar { width: 6px; height: 6px; }
*::-webkit-scrollbar-track { background: transparent; }
*::-webkit-scrollbar-thumb {
    background: rgba(255,255,255,0.08);
    border-radius: 3px;
    transition: background 0.2s;
}
*::-webkit-scrollbar-thumb:hover { background: rgba(255,255,255,0.18); }

/* Two-pane layout */
.app {
    display: grid;
    grid-template-columns: 360px 1fr;
    height: 100vh;
    height: 100dvh;
    gap: 0;
}

/* === SIDEBAR === */
.sidebar {
    background: rgba(23, 33, 43, 0.55);
    backdrop-filter: blur(40px) saturate(180%);
    -webkit-backdrop-filter: blur(40px) saturate(180%);
    border-right: 1px solid rgba(255,255,255,0.06);
    display: flex;
    flex-direction: column;
    overflow: hidden;
    box-shadow:
        4px 0 30px rgba(0,0,0,0.4),
        inset 1px 0 0 rgba(255,255,255,0.04);
}
.sidebar-header {
    padding: 18px 18px 14px;
    border-bottom: 1px solid rgba(255,255,255,0.05);
    background: linear-gradient(180deg, rgba(255,255,255,0.04), transparent);
    position: relative;
}
.sidebar-header::after {
    content: "";
    position: absolute;
    left: 0; right: 0; bottom: -1px;
    height: 1px;
    background: linear-gradient(90deg, transparent, rgba(82,136,193,0.3), transparent);
}
.sidebar-title {
    font-size: 19px;
    font-weight: 700;
    margin-bottom: 6px;
    color: #fff;
    text-shadow: 0 1px 2px rgba(0,0,0,0.3);
    letter-spacing: -0.02em;
}
.sidebar-meta { font-size: 12px; color: #7d8e9c; }
.unread-pill {
    background: linear-gradient(135deg, #ef4444, #dc2626);
    color: white;
    padding: 3px 10px;
    border-radius: 12px;
    font-size: 11px;
    font-weight: 700;
    margin-left: 6px;
    box-shadow:
        0 4px 12px rgba(239,68,68,0.4),
        inset 0 1px 0 rgba(255,255,255,0.25);
    animation: pulse-red 2s ease-in-out infinite;
}
@keyframes pulse-red {
    0%, 100% { box-shadow: 0 4px 12px rgba(239,68,68,0.4), inset 0 1px 0 rgba(255,255,255,0.25); }
    50% { box-shadow: 0 4px 20px rgba(239,68,68,0.7), inset 0 1px 0 rgba(255,255,255,0.3); }
}

.search-box { padding: 10px 14px; border-bottom: 1px solid rgba(255,255,255,0.04); }
.search-box input {
    width: 100%;
    background: rgba(36,47,61,0.6);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(255,255,255,0.05);
    color: #e6edf3;
    padding: 10px 16px;
    border-radius: 20px;
    font-size: 13px;
    outline: none;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
}
.search-box input:focus {
    background: rgba(43,57,71,0.8);
    border-color: rgba(82,136,193,0.4);
    box-shadow:
        0 0 0 4px rgba(82,136,193,0.1),
        0 4px 12px rgba(0,0,0,0.2);
}
.search-box input::placeholder { color: #7d8e9c; }

.chat-list { flex: 1; overflow-y: auto; padding: 4px 0; }

.chat-item {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 11px 14px;
    cursor: pointer;
    transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1);
    position: relative;
    margin: 2px 6px;
    border-radius: 12px;
}
.chat-item:hover {
    background: rgba(255,255,255,0.04);
    transform: translateX(2px);
}
.chat-item.active {
    background: linear-gradient(135deg, rgba(82,136,193,0.25), rgba(43,82,120,0.15));
    box-shadow:
        inset 0 0 0 1px rgba(82,136,193,0.2),
        0 4px 16px rgba(82,136,193,0.15);
}
.chat-item.active::before {
    content: "";
    position: absolute;
    left: -6px;
    top: 8px;
    bottom: 8px;
    width: 3px;
    background: linear-gradient(180deg, #5288c1, #3a6da4);
    border-radius: 0 3px 3px 0;
    box-shadow: 0 0 10px rgba(82,136,193,0.6);
}
.chat-item .avatar {
    width: 50px;
    height: 50px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-weight: 600;
    font-size: 16px;
    flex-shrink: 0;
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow:
        0 4px 14px rgba(0,0,0,0.4),
        inset 0 1px 0 rgba(255,255,255,0.25);
    text-shadow: 0 1px 2px rgba(0,0,0,0.3);
    letter-spacing: -0.02em;
}
.chat-item .info { flex: 1; min-width: 0; }
.chat-item .name {
    font-weight: 600;
    color: #fff;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    font-size: 14px;
    display: flex;
    justify-content: space-between;
}
.chat-item .name .time { font-size: 11px; color: #7d8e9c; font-weight: 400; flex-shrink: 0; margin-left: 8px; }
.chat-item .preview {
    font-size: 13px;
    color: #95a3b1;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    margin-top: 2px;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.chat-item.active .preview, .chat-item.active .name { color: #fff; }
.chat-item .new-dot {
    background: linear-gradient(135deg, #5288c1, #3a6da4);
    color: white;
    min-width: 22px;
    height: 22px;
    padding: 0 6px;
    border-radius: 11px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 11px;
    font-weight: 700;
    flex-shrink: 0;
    box-shadow:
        0 4px 12px rgba(82,136,193,0.5),
        inset 0 1px 0 rgba(255,255,255,0.25);
    animation: pulse-blue 2s ease-in-out infinite;
}
@keyframes pulse-blue {
    0%, 100% { box-shadow: 0 4px 12px rgba(82,136,193,0.5), inset 0 1px 0 rgba(255,255,255,0.25); }
    50% { box-shadow: 0 4px 20px rgba(82,136,193,0.8), inset 0 1px 0 rgba(255,255,255,0.3); }
}

/* === MAIN PANEL === */
.main {
    display: flex;
    flex-direction: column;
    overflow: hidden;
    position: relative;
}
.main-header {
    padding: 14px 22px;
    background: rgba(23, 33, 43, 0.55);
    backdrop-filter: blur(40px) saturate(180%);
    -webkit-backdrop-filter: blur(40px) saturate(180%);
    border-bottom: 1px solid rgba(255,255,255,0.05);
    display: flex;
    align-items: center;
    gap: 14px;
    box-shadow:
        0 4px 20px rgba(0,0,0,0.25),
        inset 0 1px 0 rgba(255,255,255,0.04);
    z-index: 10;
}
.main-header .avatar {
    width: 44px;
    height: 44px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-weight: 600;
    font-size: 14px;
    border: 1px solid rgba(255,255,255,0.08);
    box-shadow:
        0 4px 12px rgba(0,0,0,0.4),
        inset 0 1px 0 rgba(255,255,255,0.25);
    text-shadow: 0 1px 2px rgba(0,0,0,0.3);
}
.main-header .name {
    font-weight: 600;
    font-size: 15px;
    color: #fff;
    text-shadow: 0 1px 2px rgba(0,0,0,0.3);
}
.main-header .sub { font-size: 12px; color: #7d8e9c; }

.messages {
    flex: 1;
    overflow-y: auto;
    padding: 24px 26px;
    display: flex;
    flex-direction: column;
    gap: 4px;
}

.bubble-row {
    display: flex;
    margin: 3px 0;
    animation: fadeInUp 0.35s cubic-bezier(0.4, 0, 0.2, 1);
}
.bubble-row.right { justify-content: flex-end; }
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(10px); }
    to { opacity: 1; transform: translateY(0); }
}

.bubble {
    max-width: 65%;
    padding: 10px 14px 8px;
    border-radius: 16px;
    line-height: 1.45;
    word-wrap: break-word;
    position: relative;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    border: 1px solid rgba(255,255,255,0.06);
    backdrop-filter: blur(10px);
    -webkit-backdrop-filter: blur(10px);
}
.bubble:hover {
    transform: translateY(-2px);
    border-color: rgba(255,255,255,0.12);
}
.bubble.student {
    background: linear-gradient(135deg, rgba(82,136,193,0.95), rgba(43,82,120,0.95));
    color: #fff;
    border-bottom-right-radius: 4px;
    box-shadow:
        0 6px 20px rgba(82,136,193,0.25),
        inset 0 1px 0 rgba(255,255,255,0.15);
}
.bubble.student:hover {
    box-shadow:
        0 10px 30px rgba(82,136,193,0.4),
        inset 0 1px 0 rgba(255,255,255,0.2);
}
.bubble.bot {
    background: linear-gradient(135deg, rgba(28,43,58,0.85), rgba(20,32,46,0.85));
    color: #e6edf3;
    border-bottom-left-radius: 4px;
    box-shadow:
        0 6px 20px rgba(0,0,0,0.35),
        inset 0 1px 0 rgba(255,255,255,0.06);
}
.bubble.bot:hover {
    box-shadow:
        0 10px 30px rgba(0,0,0,0.5),
        inset 0 1px 0 rgba(255,255,255,0.1);
}
.bubble.admin {
    background: linear-gradient(135deg, rgba(217,119,6,0.95), rgba(180,83,9,0.95));
    color: #fff;
    border-bottom-left-radius: 4px;
    box-shadow:
        0 6px 20px rgba(217,119,6,0.3),
        inset 0 1px 0 rgba(255,255,255,0.2);
}
.bubble.admin:hover {
    box-shadow:
        0 10px 30px rgba(217,119,6,0.45),
        inset 0 1px 0 rgba(255,255,255,0.25);
}
.bubble .meta {
    font-size: 10px;
    opacity: 0.75;
    margin-bottom: 3px;
    font-weight: 600;
    letter-spacing: 0.02em;
    text-transform: uppercase;
}
.bubble .content { font-size: 14px; }
.bubble .time {
    font-size: 10px;
    opacity: 0.7;
    margin-top: 5px;
    text-align: right;
}
.bubble .chips {
    margin-top: 8px;
    padding-top: 8px;
    border-top: 1px solid rgba(255,255,255,0.15);
    display: flex;
    flex-wrap: wrap;
    gap: 5px;
}
.bubble .chip {
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.1);
    padding: 3px 11px;
    border-radius: 12px;
    font-size: 11px;
    font-weight: 500;
    backdrop-filter: blur(10px);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.1);
}

/* === COMPOSER === */
.composer {
    background: rgba(23, 33, 43, 0.6);
    backdrop-filter: blur(40px) saturate(180%);
    -webkit-backdrop-filter: blur(40px) saturate(180%);
    padding: 14px 22px;
    border-top: 1px solid rgba(255,255,255,0.05);
    box-shadow:
        0 -4px 20px rgba(0,0,0,0.25),
        inset 0 1px 0 rgba(255,255,255,0.04);
}
.composer-form {
    display: flex;
    align-items: center;
    gap: 10px;
    position: relative;
}
.composer textarea {
    flex: 1;
    background: rgba(36,47,61,0.6);
    backdrop-filter: blur(10px);
    border: 1px solid rgba(255,255,255,0.05);
    color: #e6edf3;
    padding: 12px 18px;
    border-radius: 22px;
    font-family: inherit;
    font-size: 14px;
    resize: none;
    min-height: 22px;
    max-height: 120px;
    line-height: 1.4;
    outline: none;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
}
.composer textarea:focus {
    background: rgba(43,57,71,0.85);
    border-color: rgba(82,136,193,0.35);
    box-shadow:
        0 0 0 4px rgba(82,136,193,0.08),
        0 4px 14px rgba(0,0,0,0.2);
}
.composer textarea::placeholder { color: #7d8e9c; }

.icon-btn {
    background: linear-gradient(135deg, #5288c1, #3a6da4);
    color: white;
    border: none;
    width: 46px;
    height: 46px;
    border-radius: 50%;
    cursor: pointer;
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    box-shadow:
        0 4px 14px rgba(82,136,193,0.4),
        inset 0 1px 0 rgba(255,255,255,0.25),
        inset 0 -1px 0 rgba(0,0,0,0.1);
}
.icon-btn:hover {
    transform: translateY(-2px);
    box-shadow:
        0 8px 22px rgba(82,136,193,0.55),
        inset 0 1px 0 rgba(255,255,255,0.3);
}
.icon-btn:active {
    transform: translateY(0);
    box-shadow:
        0 2px 6px rgba(82,136,193,0.4),
        inset 0 2px 4px rgba(0,0,0,0.2);
}
.icon-btn svg { width: 20px; height: 20px; fill: white; filter: drop-shadow(0 1px 1px rgba(0,0,0,0.2)); }

.icon-btn.attach {
    background: rgba(255,255,255,0.05);
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.05);
    color: #7d8e9c;
}
.icon-btn.attach:hover {
    background: rgba(82,136,193,0.15);
    color: #5288c1;
    transform: translateY(-1px);
    box-shadow:
        0 4px 12px rgba(82,136,193,0.2),
        inset 0 1px 0 rgba(255,255,255,0.1);
}
.icon-btn.attach svg { fill: currentColor; filter: none; }

/* Attach popup menu */
.attach-menu {
    position: absolute;
    bottom: 56px;
    left: 0;
    background: rgba(28,43,58,0.85);
    backdrop-filter: blur(40px) saturate(180%);
    -webkit-backdrop-filter: blur(40px) saturate(180%);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 14px;
    padding: 6px;
    box-shadow:
        0 20px 60px rgba(0,0,0,0.6),
        inset 0 1px 0 rgba(255,255,255,0.08);
    display: none;
    flex-direction: column;
    gap: 2px;
    z-index: 20;
    min-width: 200px;
    transform-origin: bottom left;
    animation: menuPop 0.25s cubic-bezier(0.34, 1.56, 0.64, 1);
}
@keyframes menuPop {
    from { opacity: 0; transform: scale(0.85) translateY(8px); }
    to { opacity: 1; transform: scale(1) translateY(0); }
}
.attach-menu.open { display: flex; }
.attach-menu button {
    background: none;
    border: none;
    color: #e6edf3;
    padding: 10px 14px;
    text-align: left;
    cursor: pointer;
    border-radius: 10px;
    font-size: 13px;
    display: flex;
    align-items: center;
    gap: 12px;
    transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
}
.attach-menu button:hover {
    background: rgba(82,136,193,0.15);
    transform: translateX(2px);
}
.attach-menu .icon {
    width: 32px;
    height: 32px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 15px;
    box-shadow:
        0 2px 6px rgba(0,0,0,0.3),
        inset 0 1px 0 rgba(255,255,255,0.2);
}

/* Banners */
.banner {
    margin: 12px 22px 0;
    padding: 11px 16px;
    border-radius: 12px;
    font-size: 13px;
    backdrop-filter: blur(20px);
    border: 1px solid rgba(255,255,255,0.1);
    box-shadow:
        0 6px 20px rgba(0,0,0,0.3),
        inset 0 1px 0 rgba(255,255,255,0.15);
    animation: fadeInUp 0.35s cubic-bezier(0.4, 0, 0.2, 1);
}
.banner.success { background: linear-gradient(135deg, rgba(16,185,129,0.85), rgba(5,150,105,0.85)); color: #fff; }
.banner.error { background: linear-gradient(135deg, rgba(220,38,38,0.85), rgba(153,27,27,0.85)); color: #fff; }

/* Empty state */
.empty-state {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    color: #7d8e9c;
    text-align: center;
    padding: 40px;
}
.empty-state .big-icon {
    font-size: 72px;
    margin-bottom: 20px;
    opacity: 0.35;
    filter: drop-shadow(0 8px 24px rgba(82,136,193,0.3));
    animation: float 4s ease-in-out infinite;
}
@keyframes float {
    0%, 100% { transform: translateY(0); }
    50% { transform: translateY(-8px); }
}
.empty-state h2 {
    margin: 0;
    color: #e6edf3;
    font-weight: 600;
    letter-spacing: -0.02em;
    text-shadow: 0 2px 8px rgba(0,0,0,0.4);
}

/* === Mobile-specific back button (hidden by default on desktop) === */
.back-btn {
    display: none;
    align-items: center;
    justify-content: center;
    width: 38px;
    height: 38px;
    border-radius: 50%;
    color: #e6edf3;
    text-decoration: none;
    font-size: 22px;
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.06);
    transition: all 0.2s;
    flex-shrink: 0;
}
.back-btn:hover, .back-btn:active {
    background: rgba(82,136,193,0.18);
    color: #5288c1;
}

/* === Responsive — phone layout === */
@media (max-width: 800px) {
    body { overflow: hidden; }
    .app {
        grid-template-columns: 1fr;
        height: 100vh;
        height: 100dvh;
    }
    .sidebar { display: var(--sidebar-display, flex); }
    .main { display: var(--main-display, none); }
    body.chat-open { --sidebar-display: none; --main-display: flex; }

    /* Back button: visible on mobile when in a chat */
    .back-btn { display: inline-flex !important; }

    /* Sidebar header — slightly tighter */
    .sidebar-header { padding: 14px 14px 12px; }
    .sidebar-title { font-size: 17px; }

    /* Chat header sticks to top */
    .main-header {
        position: sticky;
        top: 0;
        padding: 10px 14px;
        gap: 10px;
    }
    .main-header .name { font-size: 15px; }
    .main-header .sub { font-size: 11px; }
    .main-header .avatar { width: 38px; height: 38px; font-size: 13px; }

    /* Messages — tighter padding */
    .messages { padding: 14px 12px; }
    .bubble { max-width: 85%; padding: 8px 12px 6px; font-size: 14px; }
    .bubble .content { font-size: 14px; }

    /* Composer — sticks above keyboard */
    .composer {
        padding: 10px 12px;
        padding-bottom: max(10px, env(safe-area-inset-bottom));
    }
    .composer textarea {
        font-size: 16px; /* Prevents iOS zoom on focus */
        padding: 10px 14px;
    }
    .icon-btn { width: 42px; height: 42px; }
    .icon-btn svg { width: 18px; height: 18px; }

    /* Chat-list items — slightly tighter for phone */
    .chat-item { padding: 10px 12px; margin: 1px 4px; }
    .chat-item .avatar { width: 46px; height: 46px; font-size: 15px; }
    .chat-item .name { font-size: 14px; }
    .chat-item .preview { font-size: 13px; }

    /* Attach menu — full-width-ish on mobile */
    .attach-menu {
        left: 0;
        right: auto;
        bottom: 50px;
        min-width: 220px;
    }

    /* Banners */
    .banner { margin: 8px 12px 0; padding: 9px 13px; font-size: 12px; }
}

/* ===== PENDING LEADS PANEL (sidebar) ===== */
.pending-panel {
    background: linear-gradient(135deg, rgba(245,158,11,0.18), rgba(217,119,6,0.10));
    border-bottom: 1px solid rgba(255,255,255,0.06);
    padding: 10px 12px;
    border-left: 3px solid #f59e0b;
}
.pending-panel-title {
    font-size: 11px;
    color: #f59e0b;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    margin-bottom: 8px;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.pending-lead {
    background: rgba(0,0,0,0.25);
    border: 1px solid rgba(245,158,11,0.2);
    border-radius: 10px;
    padding: 10px;
    margin-bottom: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.2);
}
.pending-lead:last-child { margin-bottom: 0; }
.pending-lead .name { font-weight: 600; color: #fff; font-size: 13px; }
.pending-lead .meta { font-size: 11px; color: #95a3b1; margin: 2px 0 6px; }
.pending-lead .actions { display: flex; gap: 6px; }
.pending-lead .btn {
    flex: 1;
    padding: 7px 8px;
    border-radius: 8px;
    border: none;
    font-size: 11px;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.15s;
}
.pending-lead .btn-open {
    background: rgba(82,136,193,0.2);
    color: #5288c1;
    border: 1px solid rgba(82,136,193,0.3);
}
.pending-lead .btn-open:hover { background: rgba(82,136,193,0.3); }
.pending-lead .btn-called {
    background: linear-gradient(135deg, #10b981, #059669);
    color: white;
    box-shadow: 0 2px 8px rgba(16,185,129,0.3);
}
.pending-lead .btn-called:hover { transform: translateY(-1px); }

/* ===== MARK CALLED MODAL ===== */
.modal-backdrop {
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.6);
    backdrop-filter: blur(4px);
    z-index: 100;
    display: none;
    align-items: center;
    justify-content: center;
    padding: 16px;
}
.modal-backdrop.open { display: flex; animation: fadeInUp 0.25s; }
.modal-card {
    background: rgba(28,43,58,0.95);
    backdrop-filter: blur(40px) saturate(180%);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 22px;
    max-width: 520px;
    width: 100%;
    max-height: 90vh;
    overflow-y: auto;
    box-shadow: 0 20px 60px rgba(0,0,0,0.6);
}
.modal-card h2 { margin: 0 0 4px; font-size: 18px; color: #fff; }
.modal-card .modal-sub { font-size: 12px; color: #95a3b1; margin-bottom: 14px; }
.modal-card label {
    display: block;
    font-size: 11px;
    color: #95a3b1;
    margin: 12px 0 5px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.04em;
}
.modal-card input, .modal-card select, .modal-card textarea {
    width: 100%;
    background: rgba(36,47,61,0.6);
    border: 1px solid rgba(255,255,255,0.06);
    color: #e6edf3;
    padding: 10px 14px;
    border-radius: 10px;
    font-size: 13px;
    font-family: inherit;
    outline: none;
    transition: all 0.2s;
}
.modal-card textarea { min-height: 70px; resize: vertical; }
.modal-card input:focus, .modal-card select:focus, .modal-card textarea:focus {
    background: rgba(43,57,71,0.85);
    border-color: rgba(82,136,193,0.4);
}
.modal-card .csv-hint {
    font-size: 11px;
    color: #95a3b1;
    background: rgba(82,136,193,0.08);
    padding: 8px 12px;
    border-radius: 8px;
    margin-top: 8px;
    border-left: 3px solid #5288c1;
    line-height: 1.5;
}
.modal-card .csv-hint code {
    background: rgba(0,0,0,0.3);
    padding: 1px 6px;
    border-radius: 4px;
    font-size: 11px;
    color: #5288c1;
}
.modal-card .modal-actions {
    display: flex;
    gap: 10px;
    margin-top: 18px;
}
.modal-card .modal-actions button {
    flex: 1;
    padding: 12px;
    border-radius: 10px;
    border: none;
    font-weight: 600;
    cursor: pointer;
    font-size: 13px;
    transition: all 0.15s;
}
.modal-card .btn-cancel {
    background: rgba(255,255,255,0.06);
    color: #e6edf3;
    border: 1px solid rgba(255,255,255,0.08);
}
.modal-card .btn-submit {
    background: linear-gradient(135deg, #10b981, #059669);
    color: white;
    box-shadow: 0 4px 14px rgba(16,185,129,0.4);
}
.modal-card .btn-submit:hover { transform: translateY(-1px); }
"""


def _avatar_html(name: str | None, phone: str, size: int = 50) -> str:
    """Renders a colored circular avatar with initials."""
    if name and name.strip():
        parts = name.strip().split()
        initials = (parts[0][0] + (parts[-1][0] if len(parts) > 1 else "")).upper()
    else:
        initials = phone[-2:] if len(phone) >= 2 else phone
    palette = ["#8b5cf6", "#10b981", "#f59e0b", "#ef4444", "#3b82f6", "#ec4899", "#14b8a6"]
    color = palette[sum(ord(c) for c in phone) % len(palette)]
    color2 = palette[(sum(ord(c) for c in phone) + 3) % len(palette)]
    return (
        f'<div class="avatar" style="width:{size}px;height:{size}px;background:linear-gradient(135deg,{color},{color2})">'
        f'{initials}</div>'
    )


def _render_pending_leads_panel() -> str:
    """Renders the orange 'pending calls' panel that lives at the top of the sidebar."""
    leads = database.get_pending_lead_reminders()

    if not leads:
        # Show empty state so admin knows the panel exists
        return """
        <div class="pending-panel" style="opacity:0.7">
            <div class="pending-panel-title">
                <span>📞 Pending Calls (0)</span>
            </div>
            <div style="padding:8px 4px;color:#95a3b1;font-size:12px;text-align:center">
                No leads waiting. New students will appear here automatically.
            </div>
        </div>
        """

    items = []
    for lead in leads[:5]:  # Show max 5 to avoid crowding
        sender = lead["sender_id"]
        naam = (lead.get("naam") or "Unknown").replace("<", "&lt;").replace(">", "&gt;")
        cls = (lead.get("class_label") or "Class TBD").replace("<", "&lt;").replace(">", "&gt;")
        src = (lead.get("source") or "WhatsApp").replace("<", "&lt;").replace(">", "&gt;")
        # Compute waiting time
        created = lead.get("created_at") or ""
        waited = ""
        if created:
            try:
                dt = datetime.strptime(created.split(".")[0], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                mins = int((datetime.now(timezone.utc) - dt).total_seconds() / 60)
                if mins < 1:
                    waited = "just now"
                elif mins < 60:
                    waited = f"{mins} min ago"
                else:
                    waited = f"{mins // 60}h {mins % 60}m ago"
            except Exception:
                pass

        # Escape values for safe injection into onclick
        safe_naam = naam.replace("'", "&#39;")
        safe_cls = cls.replace("'", "&#39;")

        items.append(f"""
        <div class="pending-lead">
            <div class="name">{naam}</div>
            <div class="meta">{cls} · {src} · {waited}</div>
            <div class="actions">
                <button class="btn btn-open" onclick="window.location='/admin?chat={sender}'">💬 Open</button>
                <button class="btn btn-called" onclick="openCalledModal('{sender}', '{safe_naam}', '{safe_cls}', '+{sender}')">✓ Mark Called</button>
            </div>
        </div>
        """)

    extra_count = len(leads) - 5
    extra_msg = f' · <span style="color:#fff">+{extra_count} more</span>' if extra_count > 0 else ''

    return f"""
    <div class="pending-panel">
        <div class="pending-panel-title">
            <span>📞 Pending Calls ({len(leads)}){extra_msg}</span>
        </div>
        {''.join(items)}
    </div>
    """


def _render_chat_panel(sender_id: str, sent: str, error: str) -> str:
    """Renders the right-hand chat panel for a selected conversation."""
    if not sender_id:
        return """
        <div class="empty-state">
            <div class="big-icon">💬</div>
            <h2 style="margin:0;color:#e6edf3">Select a chat to start messaging</h2>
            <div style="margin-top:8px;font-size:13px">Pick a student from the sidebar to view the conversation.</div>
        </div>
        """

    # Mark as read since admin is now viewing this chat
    database.mark_chat_read(sender_id)

    messages = database.get_full_conversation(sender_id)
    conversations = database.get_all_conversations()
    chat_info = next((c for c in conversations if c["sender_id"] == sender_id), {})
    display_name = chat_info.get("display_name") or "Unknown"

    bubbles = []
    for msg in messages:
        is_user = msg["role"] == "user"
        is_admin_msg = msg["content"].startswith("[ADMIN] ")
        raw = msg["content"][8:] if is_admin_msg else msg["content"]

        if is_user:
            row_class, bubble_class, label = "right", "student", "Student"
            clean_text, btns = raw, []
        elif is_admin_msg:
            row_class, bubble_class, label = "", "admin", "You (manual)"
            clean_text, btns = raw, []
        else:
            row_class, bubble_class, label = "", "bot", "Bot"
            clean_text, btns = _clean_message_text(raw)

        content = clean_text.replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        chips_html = ""
        if btns:
            chips_html = '<div class="chips">' + "".join(
                f'<span class="chip">{b[:40].replace("<","&lt;").replace(">","&gt;")}</span>'
                for b in btns
            ) + "</div>"

        ts = to_ist(msg['timestamp'], "%d %b · %I:%M %p")
        bubbles.append(f"""
        <div class="bubble-row {row_class}">
            <div class="bubble {bubble_class}">
                <div class="meta">{label}</div>
                <div class="content">{content}</div>
                {chips_html}
                <div class="time">{ts}</div>
            </div>
        </div>
        """)

    bubbles_html = "".join(bubbles) or '<div style="text-align:center;color:#7d8e9c;padding:40px">No messages yet.</div>'

    banner = ""
    if sent == "1":
        banner = '<div class="banner success">✓ Message sent successfully</div>'
    elif error:
        err_clean = error.replace("<", "&lt;").replace(">", "&gt;")
        banner = f'<div class="banner error">✗ {err_clean}</div>'

    return f"""
    <div class="main-header">
        <a href="/admin" class="back-btn" aria-label="Back to chat list">←</a>
        {_avatar_html(display_name, sender_id, 42)}
        <div style="flex:1;min-width:0;overflow:hidden">
            <div class="name" style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{display_name}</div>
            <div class="sub" style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap">+{sender_id} · {len(messages)} messages</div>
        </div>
        <form method="POST" action="/admin/chat/{sender_id}/delete" onsubmit="return confirm('Delete entire chat with {display_name}? This cannot be undone.')" style="margin:0">
            <button type="submit" title="Delete chat" aria-label="Delete chat" style="background:rgba(220,38,38,0.15);border:1px solid rgba(220,38,38,0.3);color:#ef4444;width:38px;height:38px;border-radius:50%;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all 0.2s">
                <svg width="18" height="18" viewBox="0 0 24 24" fill="currentColor"><path d="M6 19c0 1.1.9 2 2 2h8c1.1 0 2-.9 2-2V7H6v12zM19 4h-3.5l-1-1h-5l-1 1H5v2h14V4z"/></svg>
            </button>
        </form>
    </div>
    {banner}
    <div class="messages" id="messages">
        {bubbles_html}
    </div>
    <div class="composer">
        <form class="composer-form" method="POST" action="/admin/chat/{sender_id}/send" id="text-form">
            <button type="button" class="icon-btn attach" id="attach-toggle" title="Attach" aria-label="Attach">
                <svg viewBox="0 0 24 24"><path d="M16.5 6.5v10.5a4.5 4.5 0 1 1-9 0V5a3 3 0 1 1 6 0v11.5a1.5 1.5 0 1 1-3 0V6h-1.5v10.5a3 3 0 1 0 6 0V5a4.5 4.5 0 1 0-9 0v12a6 6 0 1 0 12 0V6.5z"/></svg>
            </button>
            <div class="attach-menu" id="attach-menu">
                <button type="button" onclick="pickFile('image/*')"><span class="icon" style="background:#3b82f6">🖼️</span> Photo</button>
                <button type="button" onclick="pickFile('video/*')"><span class="icon" style="background:#ef4444">🎬</span> Video</button>
                <button type="button" onclick="pickFile('audio/*')"><span class="icon" style="background:#10b981">🎵</span> Audio</button>
                <button type="button" onclick="pickFile('application/pdf,.doc,.docx,.xls,.xlsx,.ppt,.pptx,.txt')"><span class="icon" style="background:#f59e0b">📄</span> Document</button>
            </div>
            <textarea name="message" placeholder="Message..." required maxlength="4000" id="msg-input"></textarea>
            <button type="submit" class="icon-btn" title="Send" aria-label="Send">
                <svg viewBox="0 0 24 24"><path d="M2.01 21L23 12 2.01 3 2 10l15 2-15 2z"/></svg>
            </button>
        </form>
        <form id="media-form" method="POST" action="/admin/chat/{sender_id}/send-media" enctype="multipart/form-data" style="display:none">
            <input type="file" name="file" id="media-file" />
            <input type="hidden" name="caption" id="media-caption" />
        </form>
    </div>
    """


@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(
    chat: str = "",
    sent: str = "",
    error: str = "",
    user: dict = Depends(verify_user),
):
    """Telegram-style dashboard: sidebar with conversations + main chat panel."""
    conversations = database.get_all_conversations()
    unread_count = sum(1 for c in conversations if c.get("unread"))

    # Sidebar items
    items = []
    for c in conversations:
        is_active = c["sender_id"] == chat
        raw_msg = c["last_message"] or ""
        clean_text, _btns = _clean_message_text(raw_msg)
        preview = clean_text[:50].replace("<", "&lt;").replace(">", "&gt;")
        if len(clean_text) > 50:
            preview += "…"
        prefix = "🤖 " if c["last_role"] == "assistant" else ""
        if c["last_role"] == "assistant" and raw_msg.startswith("[ADMIN] "):
            prefix = "✓ "

        time_str = to_ist(c["last_active"], "%I:%M %p")
        new_dot = f'<span class="new-dot">{"" if c.get("unread") else ""}{"!" if c.get("unread") else ""}</span>' if c.get("unread") else ''

        display_name = c.get("display_name") or "Unknown"
        items.append(f"""
        <div class="chat-item {'active' if is_active else ''}" onclick="window.location='/admin?chat={c['sender_id']}'" data-search="{display_name.lower()} {c['sender_id']}">
            {_avatar_html(display_name, c['sender_id'], 50)}
            <div class="info">
                <div class="name"><span>{display_name}</span><span class="time">{time_str}</span></div>
                <div class="preview"><span style="overflow:hidden;text-overflow:ellipsis">{prefix}{preview}</span>{new_dot}</div>
            </div>
        </div>
        """)

    items_html = "".join(items) or '<div style="text-align:center;padding:40px;color:#7d8e9c">No conversations yet.</div>'
    unread_html = f'<span class="unread-pill">{unread_count} new</span>' if unread_count > 0 else ''

    chat_panel = _render_chat_panel(chat, sent, error)

    body_class = "chat-open" if chat else ""

    # Build page title
    title = "Kaksha Kendra Bot"
    if chat:
        matched = next((c for c in conversations if c["sender_id"] == chat), None)
        if matched:
            title += " · " + (matched.get("display_name") or f"+{chat}")

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{title}</title>
        <meta charset="utf-8"/>
        <meta name="viewport" content="width=device-width, initial-scale=1"/>
        <meta name="theme-color" content="#5288c1"/>
        <link rel="manifest" href="/manifest.json"/>
        <meta name="mobile-web-app-capable" content="yes"/>
        <meta name="apple-mobile-web-app-capable" content="yes"/>
        <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent"/>
        <meta name="apple-mobile-web-app-title" content="KK Bot"/>
        <meta http-equiv="refresh" content="60">
        <style>{_ADMIN_CSS}</style>
    </head>
    <body class="{body_class}">
        <div class="app">
            <aside class="sidebar">
                <div class="sidebar-header">
                    <div style="display:flex;justify-content:space-between;align-items:center;gap:8px">
                        <div class="sidebar-title">🎓 Kaksha Kendra Bot {unread_html}</div>
                        {('<a href="/admin/settings" title="Bot Health & Settings" style="color:#7d8e9c;text-decoration:none;width:34px;height:34px;border-radius:50%;background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.06);display:inline-flex;align-items:center;justify-content:center;font-size:16px;flex-shrink:0;transition:all 0.2s">⚙️</a>') if user.get("role") == "super_admin" else ('<span style="background:rgba(245,158,11,0.15);color:#f59e0b;padding:3px 10px;border-radius:10px;font-size:11px;font-weight:700;border:1px solid rgba(245,158,11,0.3)">TEAM</span>')}
                    </div>
                    <div class="sidebar-meta" style="margin-top:6px">{len(conversations)} chats · {user.get("username", "?")} ({user.get("role", "?").replace("_", " ")})</div>
                </div>
                {_render_pending_leads_panel()}
                <div class="search-box">
                    <input type="text" id="search" placeholder="🔍 Search students..." oninput="filterChats(this.value)" />
                </div>
                <div class="chat-list" id="chat-list">
                    {items_html}
                </div>
            </aside>
            <main class="main">
                {chat_panel}
            </main>
        </div>

        <script>
            // Live-filter sidebar
            function filterChats(q) {{
                q = q.toLowerCase().trim();
                document.querySelectorAll('.chat-item').forEach(el => {{
                    const search = el.dataset.search || '';
                    el.style.display = (!q || search.includes(q)) ? '' : 'none';
                }});
            }}

            // Auto-scroll messages to bottom
            const msgs = document.getElementById('messages');
            if (msgs) msgs.scrollTop = msgs.scrollHeight;

            // Attach menu toggle
            const attachToggle = document.getElementById('attach-toggle');
            const attachMenu = document.getElementById('attach-menu');
            if (attachToggle) {{
                attachToggle.addEventListener('click', e => {{
                    e.stopPropagation();
                    attachMenu.classList.toggle('open');
                }});
                document.addEventListener('click', () => attachMenu?.classList.remove('open'));
            }}

            // File picker -> send via media form
            window.pickFile = function(accept) {{
                const f = document.getElementById('media-file');
                f.accept = accept;
                f.onchange = function() {{
                    if (this.files[0]) {{
                        const cap = prompt('Caption (optional):', '');
                        document.getElementById('media-caption').value = cap || '';
                        document.getElementById('media-form').submit();
                    }}
                }};
                f.click();
                attachMenu.classList.remove('open');
            }};

            // Ctrl+Enter to send
            const msgInput = document.getElementById('msg-input');
            if (msgInput) {{
                msgInput.focus();
                msgInput.addEventListener('keydown', function(e) {{
                    if ((e.ctrlKey || e.metaKey) && e.key === 'Enter') {{
                        e.target.form.submit();
                    }}
                }});
            }}

            // ============================================================
            // WEB PUSH (background, even when tab closed) — service worker
            // ============================================================
            async function registerPushSW() {{
                if (!('serviceWorker' in navigator) || !('PushManager' in window)) {{
                    console.warn('Push not supported in this browser');
                    return;
                }}
                try {{
                    const reg = await navigator.serviceWorker.register('/sw.js', {{ scope: '/' }});
                    console.log('SW registered');
                    return reg;
                }} catch (e) {{
                    console.error('SW registration failed:', e);
                }}
            }}

            function urlB64ToUint8Array(b64) {{
                const padding = '='.repeat((4 - b64.length % 4) % 4);
                const s = (b64 + padding).replace(/-/g, '+').replace(/_/g, '/');
                const raw = atob(s);
                return Uint8Array.from(raw, c => c.charCodeAt(0));
            }}

            window.subscribeToPush = async function() {{
                const reg = await registerPushSW();
                if (!reg) return alert('Browser does not support push notifications');
                const perm = await Notification.requestPermission();
                if (perm !== 'granted') return alert('Permission denied — enable notifications in browser settings');

                const keyRes = await fetch('/admin/push/vapid-key');
                const {{ publicKey }} = await keyRes.json();

                let sub = await reg.pushManager.getSubscription();
                if (!sub) {{
                    sub = await reg.pushManager.subscribe({{
                        userVisibleOnly: true,
                        applicationServerKey: urlB64ToUint8Array(publicKey),
                    }});
                }}

                await fetch('/admin/push/subscribe', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify(sub.toJSON()),
                }});
                alert('✅ Notifications enabled! You will be alerted even when this tab is closed.');
            }};

            // Auto-register service worker silently on every page load
            registerPushSW();

            // ============================================================
            // FOREGROUND POLLING (for when tab IS open)
            // ============================================================
            const SOUND_DATA_URI = "data:audio/wav;base64,UklGRiQEAABXQVZFZm10IBAAAAABAAEARKwAAIhYAQACABAAZGF0YQAEAAD//w==" +
                "AAD//wAA//8AAP//AAD//wAA//8AAP//AAD//wAA//8AAP//AAD//wAA//8AAP//AAD//wAA//8AAP//AAD//wAA//8AAP//AAD//wAA";
            const originalTitle = document.title;
            let unseenCount = 0;

            // Step 1 — Ask permission on first interaction
            function ensurePermission() {{
                if ('Notification' in window && Notification.permission === 'default') {{
                    Notification.requestPermission();
                }}
            }}
            document.addEventListener('click', ensurePermission, {{ once: true }});

            // Step 2 — Poll for new students and unseen messages every 15 seconds
            async function pollNewChats() {{
                try {{
                    const res = await fetch('/admin/api/conversations', {{ credentials: 'include' }});
                    if (!res.ok) return;
                    const data = await res.json();

                    const lastSeen = JSON.parse(localStorage.getItem('lastSeenChats') || '{{}}');
                    let foundNew = 0;

                    for (const chat of data.conversations || []) {{
                        const lastActive = chat.last_active;
                        const seenAt = lastSeen[chat.sender_id];
                        if (chat.unread && (!seenAt || seenAt < lastActive)) {{
                            // New unread message we haven't notified about yet
                            const isNewLead = !seenAt;
                            const title = isNewLead
                                ? '🔔 NEW STUDENT messaged Kaksha Kendra Bot'
                                : '💬 New message from ' + (chat.display_name || '+' + chat.sender_id);
                            const body = (chat.last_message || '').slice(0, 100);
                            showNotification(title, body, chat.sender_id);
                            foundNew++;
                        }}
                        lastSeen[chat.sender_id] = lastActive;
                    }}
                    localStorage.setItem('lastSeenChats', JSON.stringify(lastSeen));

                    if (foundNew > 0) {{
                        unseenCount += foundNew;
                        flashTitle();
                    }}
                }} catch (e) {{
                    console.warn('Poll failed:', e);
                }}
            }}

            function showNotification(title, body, senderId) {{
                if ('Notification' in window && Notification.permission === 'granted') {{
                    const n = new Notification(title, {{
                        body: body,
                        icon: '/favicon.ico',
                        tag: 'kk-' + senderId,
                        requireInteraction: false,
                        silent: false,
                    }});
                    n.onclick = function() {{
                        window.focus();
                        window.location.href = '/admin?chat=' + senderId;
                        n.close();
                    }};
                    setTimeout(() => n.close(), 8000);
                }}
                // Always play sound (works even without notification permission)
                try {{
                    const audio = new Audio(SOUND_DATA_URI);
                    audio.volume = 0.6;
                    audio.play().catch(() => {{}});
                }} catch (e) {{}}
            }}

            // Flash the browser tab title when there are unseen messages
            let titleInterval = null;
            function flashTitle() {{
                if (titleInterval) return;
                let toggle = false;
                titleInterval = setInterval(() => {{
                    document.title = toggle
                        ? originalTitle
                        : '🔔 (' + unseenCount + ') New message!';
                    toggle = !toggle;
                }}, 1500);
            }}

            // Reset title flash when user comes back to the tab
            window.addEventListener('focus', () => {{
                if (titleInterval) {{
                    clearInterval(titleInterval);
                    titleInterval = null;
                    document.title = originalTitle;
                    unseenCount = 0;
                }}
            }});

            // Start polling immediately + every 15 seconds
            pollNewChats();
            setInterval(pollNewChats, 15000);

            // ============================================================
            // MARK CALLED MODAL — opens from the pending-leads panel
            // ============================================================
            window.openCalledModal = function(senderId, naam, classLabel, phone) {{
                document.getElementById('cm-sender').value = senderId;
                document.getElementById('cm-naam').value = naam || '';
                document.getElementById('cm-class').value = classLabel || '';
                document.getElementById('cm-phone').value = phone || '';
                document.getElementById('cm-status').value = 'Interested';
                document.getElementById('cm-next').value = '';
                document.getElementById('cm-notes').value = '';
                document.getElementById('called-modal').classList.add('open');
            }};

            window.closeCalledModal = function() {{
                document.getElementById('called-modal').classList.remove('open');
            }};

            window.submitCalled = async function(ev) {{
                ev.preventDefault();
                const senderId = document.getElementById('cm-sender').value;
                const payload = {{
                    naam:      document.getElementById('cm-naam').value.trim(),
                    class:     document.getElementById('cm-class').value.trim(),
                    phone:     document.getElementById('cm-phone').value.trim(),
                    status:    document.getElementById('cm-status').value.trim(),
                    next_call: document.getElementById('cm-next').value.trim(),
                    notes:     document.getElementById('cm-notes').value.trim(),
                }};
                try {{
                    const res = await fetch('/admin/api/leads/' + senderId + '/called', {{
                        method: 'POST',
                        headers: {{ 'Content-Type': 'application/json' }},
                        credentials: 'include',
                        body: JSON.stringify(payload),
                    }});
                    if (res.ok) {{
                        closeCalledModal();
                        location.reload();
                    }} else {{
                        alert('Failed: ' + res.status);
                    }}
                }} catch (e) {{
                    alert('Error: ' + e.message);
                }}
            }};

            // Allow CSV-style typing into the Status field — auto-split on commas
            document.addEventListener('DOMContentLoaded', function() {{
                const csvField = document.getElementById('cm-status');
                if (!csvField) return;
                csvField.addEventListener('paste', function(e) {{
                    const txt = (e.clipboardData || window.clipboardData).getData('text');
                    if (txt && txt.includes(',')) {{
                        e.preventDefault();
                        const parts = txt.split(',').map(s => s.trim());
                        if (parts[0]) document.getElementById('cm-naam').value = parts[0];
                        if (parts[1]) document.getElementById('cm-class').value = parts[1];
                        if (parts[2]) document.getElementById('cm-phone').value = parts[2];
                        if (parts[3]) document.getElementById('cm-status').value = parts[3];
                        if (parts[4]) document.getElementById('cm-next').value = parts[4];
                        if (parts[5]) document.getElementById('cm-notes').value = parts[5];
                    }}
                }});
            }});
        </script>

        <!-- Mark Called modal — shared across all pending-lead buttons -->
        <div id="called-modal" class="modal-backdrop" onclick="if(event.target===this) closeCalledModal()">
            <form class="modal-card" onsubmit="submitCalled(event)">
                <h2>📞 Mark Lead as Called</h2>
                <div class="modal-sub">Auto-filled fields are detected by the bot. You can also paste a comma-separated row into any field — it'll auto-split.</div>
                <input type="hidden" id="cm-sender" />

                <label>Naam (Name)</label>
                <input type="text" id="cm-naam" placeholder="Student's name" />

                <label>Class</label>
                <input type="text" id="cm-class" placeholder="Class 9 / Junior / etc." />

                <label>Phone</label>
                <input type="text" id="cm-phone" placeholder="+91..." />

                <label>Status *</label>
                <select id="cm-status">
                    <option>Interested</option>
                    <option>Callback</option>
                    <option>Not Interested</option>
                    <option>Wrong Number</option>
                    <option>Enrolled</option>
                    <option>Follow-up</option>
                </select>

                <label>Next Call</label>
                <input type="text" id="cm-next" placeholder="Kal 5 PM, Tomorrow 10 AM, etc." />

                <label>Notes</label>
                <textarea id="cm-notes" placeholder="What did the student say? Concerns, requirements..."></textarea>

                <div class="csv-hint">
                    💡 <strong>Power tip:</strong> Paste a CSV row anywhere and it auto-splits:<br>
                    <code>Pari Maheshwari, Class 9, +919808884546, Interested, Kal 5 PM, Science weak hai</code>
                </div>

                <div class="modal-actions">
                    <button type="button" class="btn-cancel" onclick="closeCalledModal()">Cancel</button>
                    <button type="submit" class="btn-submit">✓ Mark Called</button>
                </div>
            </form>
        </div>
    </body>
    </html>
    """


@app.get("/admin/chat/{sender_id}", response_class=HTMLResponse)
def admin_chat_view_redirect(sender_id: str, sent: str = "", error: str = "", user: dict = Depends(verify_user)):
    """Backwards-compat: redirect old chat URLs to new query-string format."""
    qs = f"?chat={sender_id}"
    if sent:
        qs += f"&sent={sent}"
    if error:
        qs += f"&error={error}"
    return RedirectResponse(url=f"/admin{qs}", status_code=303)


@app.post("/admin/chat/{sender_id}/send")
def admin_send_message(sender_id: str, message: str = Form(...), user: dict = Depends(verify_user)):
    """Sends a manual message from admin to a student via WhatsApp."""
    message = message.strip()
    if not message:
        return RedirectResponse(url=f"/admin?chat={sender_id}&error=Empty+message", status_code=303)

    result = send_whatsapp_message(sender_id, message)

    if result is None:
        return RedirectResponse(
            url=f"/admin?chat={sender_id}&error=Failed+to+send+(check+logs)",
            status_code=303,
        )

    # Save to DB with [ADMIN] prefix so we can render it differently
    database.save_message(sender_id, "assistant", f"[ADMIN] {message}")

    return RedirectResponse(url=f"/admin?chat={sender_id}&sent=1", status_code=303)


@app.post("/admin/chat/{sender_id}/send-media")
async def admin_send_media(
    sender_id: str,
    file: UploadFile = File(...),
    caption: str = Form(""),
    user: dict = Depends(verify_user),
):
    """Uploads a media file to WhatsApp and sends it to the student."""
    file_bytes = await file.read()
    if not file_bytes:
        return RedirectResponse(url=f"/admin?chat={sender_id}&error=Empty+file", status_code=303)

    mime_type = file.content_type or "application/octet-stream"
    filename = file.filename or "file"

    # Determine media type for WhatsApp from MIME
    if mime_type.startswith("image/"):
        media_type = "image"
    elif mime_type.startswith("video/"):
        media_type = "video"
    elif mime_type.startswith("audio/"):
        media_type = "audio"
    else:
        media_type = "document"

    # Step 1: Upload to WhatsApp to get a media ID
    media_id = upload_media_to_whatsapp(file_bytes, filename, mime_type)
    if not media_id:
        return RedirectResponse(
            url=f"/admin?chat={sender_id}&error=Upload+failed+(check+logs)",
            status_code=303,
        )

    # Step 2: Send the media to the student
    success = send_whatsapp_media(
        sender_id,
        media_id,
        media_type,
        caption=caption.strip() or None,
        filename=filename if media_type == "document" else None,
    )
    if not success:
        return RedirectResponse(
            url=f"/admin?chat={sender_id}&error=Send+failed+(check+logs)",
            status_code=303,
        )

    # Step 3: Log it in the chat history so admin can see what was sent
    log_message = f"[ADMIN] [Sent {media_type}: {filename}]"
    if caption.strip():
        log_message += f" — {caption.strip()}"
    database.save_message(sender_id, "assistant", log_message)

    return RedirectResponse(url=f"/admin?chat={sender_id}&sent=1", status_code=303)


@app.post("/admin/chat/{sender_id}/delete")
def admin_delete_chat(sender_id: str, user: str = Depends(verify_admin)):
    """Deletes a conversation entirely (messages + contact info)."""
    deleted = database.delete_chat(sender_id)
    print(f"[Admin] Deleted chat with +{sender_id} ({deleted} messages)")
    return RedirectResponse(url=f"/admin?sent=1", status_code=303)


def _format_duration(seconds: float) -> str:
    """Formats seconds as a human-friendly uptime string."""
    seconds = int(seconds)
    days = seconds // 86400
    hours = (seconds % 86400) // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    parts = []
    if days: parts.append(f"{days}d")
    if hours: parts.append(f"{hours}h")
    if minutes: parts.append(f"{minutes}m")
    if not parts: parts.append(f"{secs}s")
    return " ".join(parts)


def _format_bytes(n: int) -> str:
    if n < 1024: return f"{n} B"
    if n < 1024 ** 2: return f"{n / 1024:.1f} KB"
    if n < 1024 ** 3: return f"{n / 1024 ** 2:.2f} MB"
    return f"{n / 1024 ** 3:.2f} GB"


def _render_groq_keys_card() -> str:
    """Renders a card showing per-key Groq status: success/rate-limit/idle."""
    keys = ai_stats.get("groq_keys", [])
    last_idx = ai_stats.get("last_groq_key_index")

    if not keys:
        return ""  # No Groq keys loaded yet — skip the card

    rows = []
    for k in keys:
        status = k.get("last_status", "idle")
        success = k.get("success", 0)
        rl = k.get("rate_limited", 0)
        fail = k.get("fail", 0)
        last_used = k.get("last_used_at")
        is_active = (k["index"] == last_idx)

        # Status pill colors
        if status == "success":
            dot, color, text = "🟢", "#10b981", "Healthy"
        elif status == "rate_limited":
            dot, color, text = "🟡", "#f59e0b", "Rate-limited"
        elif status == "error":
            dot, color, text = "🔴", "#ef4444", "Error"
        else:
            dot, color, text = "⚪", "#7d8e9c", "Idle"

        last_used_str = to_ist(last_used.replace("T", " ").split(".")[0], "%I:%M %p") if last_used else "—"
        active_marker = '<span style="background:rgba(82,136,193,0.18);color:#5288c1;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;margin-left:6px">LAST USED</span>' if is_active else ''

        rows.append(f"""
        <div style="display:flex;align-items:center;gap:10px;padding:10px 0;border-bottom:1px solid rgba(255,255,255,0.04)">
            <div style="width:34px;height:34px;border-radius:50%;background:rgba(255,255,255,0.05);display:flex;align-items:center;justify-content:center;font-weight:700;color:#fff;font-size:13px;flex-shrink:0;border:1px solid rgba(255,255,255,0.06)">#{k['index']}</div>
            <div style="flex:1;min-width:0">
                <div style="font-weight:600;font-size:13px;color:#fff;display:flex;align-items:center;flex-wrap:wrap">
                    <span>{k['label']}</span>{active_marker}
                </div>
                <div style="font-size:11px;color:#7d8e9c;font-family:monospace">{k.get('key_preview', '***')} · last: {last_used_str}</div>
            </div>
            <div style="text-align:right;flex-shrink:0">
                <div style="color:{color};font-weight:600;font-size:13px">{dot} {text}</div>
                <div style="font-size:11px;color:#7d8e9c">✓{success} · 🟡{rl} · ✗{fail}</div>
            </div>
        </div>
        """)

    return f"""
    <div class="card" style="grid-column: span 2">
        <h3>🔑 Groq API Keys ({len(keys)})</h3>
        <div style="font-size:12px;color:#95a3b1;margin-bottom:8px">
            Bot tries each key in order. When one rate-limits, the next takes over.
            ✓ success · 🟡 rate-limited · ✗ error
        </div>
        {''.join(rows)}
    </div>
    """


@app.get("/admin/settings", response_class=HTMLResponse)
def admin_settings(user: str = Depends(verify_admin)):
    """Bot health, AI provider stats, and database statistics."""
    db_stats = database.get_stats()

    # Uptime
    uptime_seconds = (datetime.now(timezone.utc) - SERVER_STARTED_AT).total_seconds()
    uptime_str = _format_duration(uptime_seconds)
    started_at_ist = to_ist(
        SERVER_STARTED_AT.strftime("%Y-%m-%d %H:%M:%S"),
        "%d %b %Y, %I:%M %p",
    )

    # AI provider key counts
    groq_keys = len(_get_groq_clients())
    has_gemini = bool(os.getenv("GEMINI_API_KEY"))
    has_cerebras = bool(os.getenv("CEREBRAS_API_KEY"))

    # Provider health (online if last success within 1 hour)
    def provider_status(last_iso: str | None) -> tuple[str, str]:
        if not last_iso:
            return ("⚪ Idle", "#7d8e9c")
        try:
            last = datetime.fromisoformat(last_iso.replace("Z", "+00:00"))
            mins_ago = (datetime.now(timezone.utc) - last).total_seconds() / 60
            if mins_ago < 60:
                return (f"🟢 Healthy ({int(mins_ago)}m ago)", "#10b981")
            if mins_ago < 1440:
                return (f"🟡 Recent ({int(mins_ago / 60)}h ago)", "#f59e0b")
            return (f"🔴 Stale ({int(mins_ago / 1440)}d ago)", "#ef4444")
        except Exception:
            return ("⚪ Unknown", "#7d8e9c")

    groq_status, groq_color = provider_status(ai_stats.get("last_groq_success_at"))
    gemini_status, gemini_color = provider_status(ai_stats.get("last_gemini_success_at"))
    cerebras_status, cerebras_color = provider_status(ai_stats.get("last_cerebras_success_at"))

    # Token cost estimate (very rough): Groq paid ~$0.59/M input, $0.79/M output
    # Free tier hits limit at 100k/day. Use 0 cost while within free tier estimate.
    tokens_used = ai_stats.get("tokens_used_estimate", 0)
    cost_usd = tokens_used / 1_000_000 * 0.7  # rough average
    cost_inr = cost_usd * 83  # approx exchange rate

    # AI usage breakdown for the bar chart
    total_ai = (
        ai_stats["groq_success"]
        + ai_stats["gemini_success"]
        + ai_stats["cerebras_success"]
    ) or 1
    groq_pct = (ai_stats["groq_success"] / total_ai) * 100
    gemini_pct = (ai_stats["gemini_success"] / total_ai) * 100
    cerebras_pct = (ai_stats["cerebras_success"] / total_ai) * 100

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Settings · Kaksha Kendra Bot</title>
        <meta charset="utf-8"/>
        <meta name="viewport" content="width=device-width, initial-scale=1"/>
        <meta name="theme-color" content="#5288c1"/>
        <link rel="manifest" href="/manifest.json"/>
        <meta http-equiv="refresh" content="30">
        <style>{_ADMIN_CSS}
            /* Override the dashboard's no-scroll body for settings page */
            html, body {{ height: auto !important; overflow: auto !important; }}
            body {{ min-height: 100vh; min-height: 100dvh; padding-bottom: 40px; }}
            .settings-wrap {{ padding: 20px; max-width: 1100px; margin: 0 auto; width: 100%; }}

            /* Sticky top bar — stays at top while scrolling */
            .top-bar {{
                position: sticky;
                top: 0;
                z-index: 50;
                background: rgba(14, 22, 33, 0.85);
                backdrop-filter: blur(30px) saturate(180%);
                -webkit-backdrop-filter: blur(30px) saturate(180%);
                margin: -20px -20px 0;
                padding: 14px 20px;
                border-bottom: 1px solid rgba(255,255,255,0.05);
            }}

            /* === MOBILE — phone-first redesign === */
            @media (max-width: 800px) {{
                body {{ padding-bottom: 60px; }}
                .settings-wrap {{ padding: 0; }}
                .top-bar {{
                    margin: 0;
                    padding: 14px 16px;
                    border-radius: 0;
                }}
                .top-bar h1 {{ font-size: 17px; font-weight: 600; }}
                .nav-link {{
                    padding: 8px 14px;
                    font-size: 12px;
                    border-radius: 18px;
                }}

                /* Force single-column on phones, even cards that span 2 */
                .settings-grid {{
                    grid-template-columns: 1fr !important;
                    gap: 12px;
                    padding: 14px;
                    margin-top: 6px !important;
                }}
                .card[style*="span"] {{ grid-column: 1 !important; }}

                /* Cards: tighter, snappier, native-app feel */
                .card {{
                    padding: 16px;
                    border-radius: 16px;
                }}
                .card:hover {{ transform: none; }}  /* No hover transform on touch */
                .card h3 {{
                    font-size: 11px;
                    margin-bottom: 10px;
                    letter-spacing: 0.08em;
                }}
                .card .big {{ font-size: 26px; line-height: 1.0; }}
                .card .small {{ font-size: 12px; }}

                /* Rows: stack values cleanly, larger touch targets */
                .row {{
                    padding: 12px 0;
                    flex-wrap: wrap;
                    gap: 4px;
                }}
                .row .label {{ font-size: 13px; }}
                .row .val {{ font-size: 13px; }}

                /* Buttons in cards stack on tiny screens */
                .card a, .card button {{
                    min-height: 44px;  /* iOS touch-target minimum */
                    width: 100%;
                    justify-content: center;
                }}
                .card div[style*="display:flex;gap:10px"] {{
                    flex-direction: column;
                }}

                /* Quick links — touchable rows */
                .card .row a {{
                    width: auto;
                    min-height: auto;
                }}

                /* Visual provider bar — slimmer */
                .bar {{ height: 6px; }}

                /* Bottom footer */
                body > div > div:last-child {{ font-size: 11px; padding: 0 16px; }}
            }}

            /* Very small phones (iPhone SE etc.) */
            @media (max-width: 380px) {{
                .top-bar h1 {{ font-size: 15px; }}
                .card {{ padding: 14px; }}
                .card .big {{ font-size: 22px; }}
            }}
            .settings-grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
                gap: 14px;
                margin-top: 18px;
            }}
            .card {{
                background: rgba(23, 33, 43, 0.55);
                backdrop-filter: blur(40px) saturate(180%);
                -webkit-backdrop-filter: blur(40px) saturate(180%);
                border: 1px solid rgba(255,255,255,0.06);
                border-radius: 14px;
                padding: 18px;
                box-shadow: 0 8px 32px rgba(0,0,0,0.3), inset 0 1px 0 rgba(255,255,255,0.04);
                transition: transform 0.25s cubic-bezier(0.4, 0, 0.2, 1);
            }}
            .card:hover {{ transform: translateY(-3px); }}
            .card h3 {{
                margin: 0 0 12px;
                font-size: 13px;
                color: #95a3b1;
                font-weight: 600;
                text-transform: uppercase;
                letter-spacing: 0.05em;
            }}
            .card .big {{ font-size: 28px; font-weight: 700; color: #fff; line-height: 1.1; }}
            .card .small {{ font-size: 13px; color: #95a3b1; margin-top: 4px; }}
            .row {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid rgba(255,255,255,0.04); }}
            .row:last-child {{ border-bottom: none; }}
            .row .label {{ color: #95a3b1; font-size: 13px; }}
            .row .val {{ color: #fff; font-weight: 600; font-size: 13px; }}
            .bar {{
                height: 8px;
                background: rgba(255,255,255,0.05);
                border-radius: 4px;
                overflow: hidden;
                display: flex;
                margin-top: 8px;
            }}
            .bar > div {{ height: 100%; transition: width 0.5s; }}
            .badge {{ display: inline-block; padding: 3px 10px; border-radius: 10px; font-size: 11px; font-weight: 700; }}
            .top-bar {{ display: flex; justify-content: space-between; align-items: center; gap: 12px; padding-bottom: 8px; }}
            .top-bar h1 {{ margin: 0; font-size: 22px; }}
            .nav-link {{
                color: #5288c1;
                background: rgba(82,136,193,0.1);
                border: 1px solid rgba(82,136,193,0.2);
                padding: 8px 14px;
                border-radius: 20px;
                text-decoration: none;
                font-size: 13px;
            }}
        </style>
    </head>
    <body>
        <div class="settings-wrap">
            <div class="top-bar">
                <h1>⚙️ Bot Health & Settings</h1>
                <a href="/admin" class="nav-link">← Back to dashboard</a>
            </div>

            <div class="settings-grid">

                <!-- ================= UPTIME ================= -->
                <div class="card">
                    <h3>🟢 Server Uptime</h3>
                    <div class="big">{uptime_str}</div>
                    <div class="small">Started: {started_at_ist} IST</div>
                </div>

                <!-- ================= AI Calls ================= -->
                <div class="card">
                    <h3>🤖 AI Calls (Since Restart)</h3>
                    <div class="big">{ai_stats["total_calls"]}</div>
                    <div class="small">{ai_stats["all_failed"]} fully failed · {total_ai} succeeded</div>
                    <div class="bar">
                        <div style="width:{groq_pct}%;background:#5288c1" title="Groq"></div>
                        <div style="width:{gemini_pct}%;background:#10b981" title="Gemini"></div>
                        <div style="width:{cerebras_pct}%;background:#f59e0b" title="Cerebras"></div>
                    </div>
                    <div class="small" style="display:flex;gap:12px;margin-top:8px">
                        <span><span style="color:#5288c1">●</span> Groq {ai_stats["groq_success"]}</span>
                        <span><span style="color:#10b981">●</span> Gemini {ai_stats["gemini_success"]}</span>
                        <span><span style="color:#f59e0b">●</span> Cerebras {ai_stats["cerebras_success"]}</span>
                    </div>
                </div>

                <!-- ================= Token Usage ================= -->
                <div class="card">
                    <h3>💰 Token Usage Estimate</h3>
                    <div class="big">{tokens_used:,}</div>
                    <div class="small">Approximate cost: ${cost_usd:.4f} (~₹{cost_inr:.2f})</div>
                    <div class="small" style="margin-top:6px;font-size:11px">Free tiers cover most usage. Cost shown is paid-tier estimate.</div>
                </div>

                <!-- ================= Provider Status ================= -->
                <div class="card" style="grid-column: span 2">
                    <h3>📡 AI Provider Status</h3>
                    <div class="row">
                        <span class="label">Groq <span class="badge" style="background:rgba(82,136,193,0.15);color:#5288c1">{groq_keys} key{"s" if groq_keys != 1 else ""}</span></span>
                        <span class="val" style="color:{groq_color}">{groq_status}</span>
                    </div>
                    <div class="row">
                        <span class="label">Gemini <span class="badge" style="background:rgba(16,185,129,0.15);color:#10b981">{"On" if has_gemini else "Off"}</span></span>
                        <span class="val" style="color:{gemini_color}">{gemini_status}</span>
                    </div>
                    <div class="row">
                        <span class="label">Cerebras <span class="badge" style="background:rgba(245,158,11,0.15);color:#f59e0b">{"On" if has_cerebras else "Off"}</span></span>
                        <span class="val" style="color:{cerebras_color}">{cerebras_status}</span>
                    </div>
                    <div class="row">
                        <span class="label">Failures (Groq · Gemini · Cerebras)</span>
                        <span class="val">{ai_stats["groq_fail"]} · {ai_stats["gemini_fail"]} · {ai_stats["cerebras_fail"]}</span>
                    </div>
                </div>

                <!-- ================= Per-Key Groq Status ================= -->
                {_render_groq_keys_card()}

                <!-- ================= Students Today ================= -->
                <div class="card">
                    <h3>📥 Students (Last 24h)</h3>
                    <div class="big">{db_stats["new_students_today"]}</div>
                    <div class="small">{db_stats["messages_today"]} total messages today</div>
                </div>

                <!-- ================= All-time stats ================= -->
                <div class="card" style="grid-column: span 2">
                    <h3>📊 Conversation Stats</h3>
                    <div class="row">
                        <span class="label">Total students (all time)</span>
                        <span class="val">{db_stats["total_students"]}</span>
                    </div>
                    <div class="row">
                        <span class="label">Active in last 7 days</span>
                        <span class="val">{db_stats["active_students_week"]}</span>
                    </div>
                    <div class="row">
                        <span class="label">Total messages stored</span>
                        <span class="val">{db_stats["total_messages"]:,}</span>
                    </div>
                    <div class="row">
                        <span class="label">Student → Bot messages</span>
                        <span class="val">{db_stats["total_user_messages"]:,}</span>
                    </div>
                    <div class="row">
                        <span class="label">Bot → Student replies</span>
                        <span class="val">{db_stats["total_bot_replies"]:,}</span>
                    </div>
                    <div class="row">
                        <span class="label">Messages this week</span>
                        <span class="val">{db_stats["messages_week"]:,}</span>
                    </div>
                    <div class="row">
                        <span class="label">Database size</span>
                        <span class="val">{_format_bytes(db_stats["db_size_bytes"])}</span>
                    </div>
                </div>

                <!-- ================= Push Notifications ================= -->
                <div class="card" style="grid-column: span 2">
                    <h3>🔔 Push Notifications (Background)</h3>
                    <p style="font-size:13px;color:#95a3b1;margin:0 0 12px">
                        Get instant alerts when new students message — even when the dashboard is closed.
                        Works on PC and mobile (install as PWA for best experience).
                    </p>
                    <div style="display:flex;gap:10px;flex-wrap:wrap">
                        <button onclick="subscribeToPush()" style="background:linear-gradient(135deg,#5288c1,#3a6da4);color:#fff;border:none;padding:10px 18px;border-radius:22px;cursor:pointer;font-weight:600;font-size:13px;box-shadow:0 4px 12px rgba(82,136,193,0.4)">
                            🔔 Enable Notifications
                        </button>
                        <button onclick="testPush()" style="background:rgba(16,185,129,0.15);color:#10b981;border:1px solid rgba(16,185,129,0.3);padding:10px 18px;border-radius:22px;cursor:pointer;font-weight:600;font-size:13px">
                            🧪 Send Test Push
                        </button>
                    </div>
                    <div style="margin-top:14px;font-size:12px;color:#7d8e9c">
                        💡 <strong>For background alerts on phone:</strong> Browser menu → "Install app" or "Add to Home Screen"
                    </div>
                </div>

                <!-- ================= Google Sheets Live Sync ================= -->
                <div class="card" style="grid-column: span 2">
                    <h3>📊 Google Sheets Live Sync {('<span class="badge" style="background:rgba(16,185,129,0.15);color:#10b981">ON</span>' if os.getenv('GOOGLE_SHEETS_WEBHOOK') else '<span class="badge" style="background:rgba(125,142,156,0.15);color:#7d8e9c">OFF</span>')}</h3>
                    <p style="font-size:13px;color:#95a3b1;margin:0 0 10px">
                        Auto-sync every student message to a Google Sheet for backup, sharing with team, or analytics.
                    </p>
                    <details style="margin-top:8px">
                        <summary style="cursor:pointer;color:#5288c1;font-size:13px;font-weight:600;padding:6px 0">📖 Setup instructions (one-time, ~5 min)</summary>
                        <div style="margin-top:10px;padding:14px;background:rgba(0,0,0,0.2);border-radius:10px;border:1px solid rgba(255,255,255,0.05);font-size:12px;line-height:1.6;color:#c5cdd5">
                            <strong style="color:#fff">Step 1 — Create a Google Sheet</strong><br>
                            Go to <a href="https://sheets.new" target="_blank" style="color:#5288c1">sheets.new</a>. Add these <strong>7 column headers</strong> in row 1 (CRM-style):<br>
                            <code style="background:#0e1621;padding:2px 6px;border-radius:4px">Naam | Class | Phone | Source | Status | Next Call | Notes</code>
                            <br><br>
                            <strong style="color:#fff">Step 2 — Add Apps Script</strong><br>
                            In the sheet, click <code>Extensions → Apps Script</code>. Replace all code with the script below and save.
                            <br><br>
                            <strong style="color:#fff">Step 3 — Deploy as Web App</strong><br>
                            Apps Script → <code>Deploy → New deployment → Web app</code>.<br>
                            • Execute as: <code>Me</code> &nbsp; • Who has access: <code>Anyone</code><br>
                            Copy the URL (ends with <code>/exec</code>).
                            <br><br>
                            <strong style="color:#fff">Step 4 — Add to Railway</strong><br>
                            Railway → Variables → <code>GOOGLE_SHEETS_WEBHOOK</code> = paste URL → save.
                            <br><br>
                            <strong style="color:#fff">Auto-detected fields:</strong>
                            <ul style="margin:4px 0 0 18px;padding:0">
                                <li><strong>Naam</strong> — from WhatsApp profile</li>
                                <li><strong>Class</strong> — detected from message ("Class 9", "10th", "Pre-Primary", etc.)</li>
                                <li><strong>Phone</strong> — student's number</li>
                                <li><strong>Source</strong> — Meta Ad / Referral / Website / WhatsApp / Instagram</li>
                                <li><strong>Status</strong> — starts as "New" for new leads</li>
                                <li><strong>Next Call & Notes</strong> — you fill these manually; the bot will <em>NEVER overwrite them</em></li>
                            </ul>
                            <br>
                            <strong style="color:#fff">Apps Script code (copy and paste):</strong>
                            <pre style="background:#0a1320;padding:12px;border-radius:6px;font-size:11px;overflow-x:auto;margin-top:6px;color:#a5b4c4">// Cols: A=Naam B=Class C=Phone D=Source E=Status F=Next Call G=Notes
function doPost(e) {{
  var sheet = SpreadsheetApp.getActiveSpreadsheet().getActiveSheet();
  var d = JSON.parse(e.postData.contents);
  var lastRow = sheet.getLastRow();
  var phones = lastRow >= 2 ? sheet.getRange(2, 3, lastRow - 1, 1).getValues().map(function(r){{return r[0];}}) : [];
  var rowIdx = phones.indexOf(d.phone);

  if (rowIdx === -1) {{
    // New lead — write all auto fields, leave Next Call + Notes blank for admin
    sheet.appendRow([
      d.naam || "",
      d.class || "",
      d.phone || "",
      d.source || "WhatsApp",
      d.status || "New",
      "",  // Next Call (admin fills)
      ""   // Notes (admin fills)
    ]);
  }} else {{
    // Existing lead — update auto fields ONLY, preserve admin-edited columns
    var sheetRow = rowIdx + 2;
    if (d.naam)   sheet.getRange(sheetRow, 1).setValue(d.naam);
    if (d.class)  sheet.getRange(sheetRow, 2).setValue(d.class);
    sheet.getRange(sheetRow, 3).setValue(d.phone);
    // Source / Status / Next Call / Notes are NEVER overwritten on follow-ups
  }}

  return ContentService
    .createTextOutput(JSON.stringify({{ok: true}}))
    .setMimeType(ContentService.MimeType.JSON);
}}</pre>
                        </div>
                    </details>
                </div>

                <!-- ================= Export Data ================= -->
                <div class="card">
                    <h3>📥 Export Data</h3>
                    <a href="/admin/export/csv" download style="display:flex;align-items:center;gap:10px;padding:12px;background:rgba(16,185,129,0.1);border:1px solid rgba(16,185,129,0.3);border-radius:10px;color:#10b981;text-decoration:none;font-weight:600;font-size:13px;margin-bottom:8px;transition:all 0.2s">
                        <span style="font-size:18px">📄</span>
                        <div style="flex:1">
                            <div>Download CSV</div>
                            <div style="font-size:11px;color:#7d8e9c;font-weight:400">Plain text, opens in Excel</div>
                        </div>
                        <span>↓</span>
                    </a>
                    <a href="/admin/export/xlsx" download style="display:flex;align-items:center;gap:10px;padding:12px;background:rgba(82,136,193,0.1);border:1px solid rgba(82,136,193,0.3);border-radius:10px;color:#5288c1;text-decoration:none;font-weight:600;font-size:13px;transition:all 0.2s">
                        <span style="font-size:18px">📊</span>
                        <div style="flex:1">
                            <div>Download Excel (.xlsx)</div>
                            <div style="font-size:11px;color:#7d8e9c;font-weight:400">Formatted, 2 sheets (Summary + Messages)</div>
                        </div>
                        <span>↓</span>
                    </a>
                </div>

                <!-- ================= Quick Links ================= -->
                <div class="card">
                    <h3>🔗 Quick Links</h3>
                    <div class="row">
                        <a href="https://railway.com" target="_blank" class="label" style="color:#5288c1">Railway Dashboard ↗</a>
                    </div>
                    <div class="row">
                        <a href="https://console.groq.com/settings/usage" target="_blank" class="label" style="color:#5288c1">Groq Usage ↗</a>
                    </div>
                    <div class="row">
                        <a href="https://aistudio.google.com/app/apikey" target="_blank" class="label" style="color:#5288c1">Gemini API Keys ↗</a>
                    </div>
                    <div class="row">
                        <a href="https://cloud.cerebras.ai/platform/inference/limits" target="_blank" class="label" style="color:#5288c1">Cerebras Limits ↗</a>
                    </div>
                </div>

            </div>

            <div style="margin-top: 16px; text-align: center; color: #7d8e9c; font-size: 12px">
                Auto-refreshing every 30s · Stats reset on server restart (last restart: {started_at_ist} IST)
            </div>
        </div>

        <script>
            // Service worker registration for push
            async function registerSW() {{
                if (!('serviceWorker' in navigator) || !('PushManager' in window)) return null;
                try {{ return await navigator.serviceWorker.register('/sw.js', {{ scope: '/' }}); }}
                catch (e) {{ console.error(e); return null; }}
            }}
            registerSW();

            function urlB64ToUint8Array(b64) {{
                const padding = '='.repeat((4 - b64.length % 4) % 4);
                const s = (b64 + padding).replace(/-/g, '+').replace(/_/g, '/');
                const raw = atob(s);
                return Uint8Array.from(raw, c => c.charCodeAt(0));
            }}

            window.subscribeToPush = async function() {{
                const reg = await registerSW();
                if (!reg) return alert('Browser does not support push notifications');
                const perm = await Notification.requestPermission();
                if (perm !== 'granted') return alert('Permission denied. Enable in browser settings.');

                const keyRes = await fetch('/admin/push/vapid-key');
                const {{ publicKey }} = await keyRes.json();

                let sub = await reg.pushManager.getSubscription();
                if (!sub) {{
                    sub = await reg.pushManager.subscribe({{
                        userVisibleOnly: true,
                        applicationServerKey: urlB64ToUint8Array(publicKey),
                    }});
                }}

                const res = await fetch('/admin/push/subscribe', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify(sub.toJSON()),
                }});
                if (res.ok) alert('✅ Notifications enabled! You will be alerted even when this tab is closed.');
                else alert('❌ Failed to subscribe. Check console.');
            }};

            window.testPush = async function() {{
                const res = await fetch('/admin/push/test', {{ method: 'POST' }});
                if (res.ok) alert('✅ Test push sent — check for the notification!');
                else alert('❌ Test failed. Check Railway logs.');
            }};
        </script>
    </body>
    </html>
    """


@app.get("/admin/export/csv")
def admin_export_csv(user: str = Depends(verify_admin)):
    """Exports all chats as a CSV file (one row per message)."""
    rows = database.get_all_messages_with_contact()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["ID", "Phone", "Name", "Role", "Message", "Timestamp (IST)"])
    for r in rows:
        writer.writerow([
            r["id"],
            "+" + r["sender_id"],
            r["display_name"],
            r["role"],
            r["content"],
            to_ist(r["timestamp"], "%Y-%m-%d %H:%M:%S"),
        ])

    buf.seek(0)
    filename = f"kakshakendra_chats_{datetime.now(IST).strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/admin/export/xlsx")
def admin_export_xlsx(user: str = Depends(verify_admin)):
    """
    Exports all chats as an Excel file with two sheets:
    1. 'Summary' — one row per student with stats
    2. 'Messages' — every message in detail
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed; redeploy needed")

    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws_sum = wb.active
    ws_sum.title = "Summary"

    summary_headers = ["Phone", "Name", "Total Messages", "Last Active (IST)", "Last Message", "Status"]
    ws_sum.append(summary_headers)

    conversations = database.get_all_conversations()
    for c in conversations:
        last_msg = (c["last_message"] or "")[:200]
        status = "🔔 NEW" if c.get("unread") else "Read"
        ws_sum.append([
            "+" + c["sender_id"],
            c.get("display_name") or "",
            c["message_count"],
            to_ist(c["last_active"], "%Y-%m-%d %H:%M"),
            last_msg,
            status,
        ])

    # Style summary sheet
    header_fill = PatternFill(start_color="2B5278", end_color="2B5278", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_num, _ in enumerate(summary_headers, 1):
        cell = ws_sum.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-width columns
    column_widths_sum = [16, 22, 14, 22, 60, 10]
    for i, w in enumerate(column_widths_sum, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    ws_sum.freeze_panes = "A2"

    # ---------- Sheet 2: All Messages ----------
    ws_msg = wb.create_sheet(title="Messages")
    msg_headers = ["ID", "Phone", "Name", "Role", "Message", "Timestamp (IST)"]
    ws_msg.append(msg_headers)

    rows = database.get_all_messages_with_contact()
    for r in rows:
        ws_msg.append([
            r["id"],
            "+" + r["sender_id"],
            r["display_name"],
            r["role"],
            r["content"],
            to_ist(r["timestamp"], "%Y-%m-%d %H:%M:%S"),
        ])

    for col_num, _ in enumerate(msg_headers, 1):
        cell = ws_msg.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    column_widths_msg = [8, 16, 22, 12, 80, 22]
    for i, w in enumerate(column_widths_msg, 1):
        ws_msg.column_dimensions[get_column_letter(i)].width = w

    # Wrap text in message column
    for row in ws_msg.iter_rows(min_row=2, max_col=5, min_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_msg.freeze_panes = "A2"

    # ---------- Write to bytes ----------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kakshakendra_chats_{datetime.now(IST).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================
# Web Push Endpoints — service worker, manifest, subscribe API
# ============================================================
@app.get("/sw.js")
def service_worker():
    """Returns the service worker JS that handles push events."""
    js = """
self.addEventListener('install', e => self.skipWaiting());
self.addEventListener('activate', e => e.waitUntil(self.clients.claim()));

self.addEventListener('push', event => {
    let data = { title: 'Kaksha Kendra Bot', body: 'New notification' };
    try { if (event.data) data = event.data.json(); } catch (e) {}
    event.waitUntil(
        self.registration.showNotification(data.title, {
            body: data.body || '',
            icon: data.icon || '/favicon.ico',
            badge: data.icon || '/favicon.ico',
            data: { url: data.url || '/admin' },
            tag: data.tag || 'kk-notification',
            renotify: true,
            requireInteraction: false,
            vibrate: [200, 100, 200]
        })
    );
});

self.addEventListener('notificationclick', event => {
    event.notification.close();
    const targetUrl = (event.notification.data && event.notification.data.url) || '/admin';
    event.waitUntil(
        self.clients.matchAll({ type: 'window', includeUncontrolled: true }).then(list => {
            for (const client of list) {
                if (client.url.includes('/admin') && 'focus' in client) {
                    client.navigate(targetUrl);
                    return client.focus();
                }
            }
            if (self.clients.openWindow) return self.clients.openWindow(targetUrl);
        })
    );
});
"""
    return Response(content=js, media_type="application/javascript")


@app.get("/manifest.json")
def manifest():
    """PWA manifest so admins can 'Install app' to home screen."""
    return JSONResponse({
        "name": "Kaksha Kendra Bot",
        "short_name": "KK Bot",
        "description": "Admin dashboard for Kaksha Kendra WhatsApp bot",
        "start_url": "/admin",
        "scope": "/",
        "display": "standalone",
        "background_color": "#0e1621",
        "theme_color": "#5288c1",
        "orientation": "portrait",
        "icons": [
            {
                "src": "/favicon.ico",
                "sizes": "any",
                "type": "image/x-icon",
                "purpose": "any",
            },
        ],
    })


@app.get("/admin/push/vapid-key")
def push_vapid_key(user: dict = Depends(verify_user)):
    """Returns the public VAPID key — browser uses it to subscribe."""
    return {"publicKey": VAPID_PUBLIC, "role": user["role"]}


@app.post("/admin/push/subscribe")
async def push_subscribe(request: Request, user: dict = Depends(verify_user)):
    """Saves a browser push subscription tagged with the user's role."""
    body = await request.json()
    endpoint = body.get("endpoint", "")
    keys = body.get("keys", {}) or {}
    p256dh = keys.get("p256dh", "")
    auth = keys.get("auth", "")
    if not (endpoint and p256dh and auth):
        raise HTTPException(status_code=400, detail="Missing subscription fields")
    database.add_push_subscription(endpoint, p256dh, auth, role=user["role"])
    return {"status": "subscribed", "role": user["role"]}


@app.post("/admin/push/unsubscribe")
async def push_unsubscribe(request: Request, user: dict = Depends(verify_user)):
    """Removes a browser push subscription."""
    body = await request.json()
    endpoint = body.get("endpoint", "")
    if endpoint:
        database.remove_push_subscription(endpoint)
    return {"status": "unsubscribed"}


@app.post("/admin/push/test")
def push_test(user: str = Depends(verify_admin)):
    """Send a test push to verify everything works."""
    send_web_push(
        title="🎓 Test Notification",
        body="If you see this, push notifications work!",
        url="/admin",
        tag="kk-test",
    )
    return {"status": "sent"}


@app.get("/call/{number}", response_class=HTMLResponse)
def call_redirect(number: str):
    """
    Public 'tap-to-call' redirect page.
    Used as a CTA URL button target — when student taps the button in WhatsApp,
    it opens this page in their browser, which immediately triggers the dialer
    via tel:+<number>.
    """
    # Sanitize: keep digits only
    clean = "".join(c for c in number if c.isdigit())
    return f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8"/>
    <meta name="viewport" content="width=device-width, initial-scale=1"/>
    <title>Calling Kaksha Kendra...</title>
    <meta http-equiv="refresh" content="0; url=tel:+{clean}"/>
    <style>
        body {{
            font-family: -apple-system, sans-serif;
            background: linear-gradient(135deg, #0a1320, #14202d);
            color: #e6edf3;
            margin: 0;
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            text-align: center;
            padding: 20px;
        }}
        .ring {{
            font-size: 80px;
            animation: ring 0.6s ease-in-out infinite;
        }}
        @keyframes ring {{
            0%, 100% {{ transform: rotate(-15deg); }}
            50% {{ transform: rotate(15deg); }}
        }}
        h1 {{ margin: 20px 0 8px; font-size: 22px; }}
        p {{ color: #95a3b1; margin: 4px 0; }}
        a {{
            display: inline-block;
            margin-top: 24px;
            background: linear-gradient(135deg, #5288c1, #3a6da4);
            color: white;
            padding: 14px 28px;
            border-radius: 30px;
            text-decoration: none;
            font-weight: 600;
            box-shadow: 0 6px 20px rgba(82,136,193,0.4);
        }}
    </style>
</head>
<body>
    <div class="ring">📞</div>
    <h1>Calling Kaksha Kendra</h1>
    <p>+{clean[:2]} {clean[2:7]} {clean[7:]}</p>
    <p style="font-size:13px">If your dialer didn't open automatically, tap below:</p>
    <a href="tel:+{clean}">📞 Call Now</a>
    <script>
        // Belt-and-suspenders: also try via JS in case meta-refresh is blocked
        setTimeout(function() {{
            window.location.href = "tel:+{clean}";
        }}, 100);
    </script>
</body>
</html>"""


@app.get("/admin/api/conversations")
def admin_api_conversations(user: dict = Depends(verify_user)):
    """JSON endpoint listing all conversations."""
    return {"conversations": database.get_all_conversations()}


# ============================================================
# Lead Reminder API — used by team and admin to track who to call
# ============================================================

@app.get("/admin/api/leads/pending")
def api_leads_pending(user: dict = Depends(verify_user)):
    """Returns the list of leads still waiting for a call."""
    return {"pending": database.get_pending_lead_reminders()}


@app.post("/admin/api/leads/{sender_id}/called")
async def api_lead_called(
    sender_id: str,
    request: Request,
    user: dict = Depends(verify_user),
):
    """
    Marks a lead as called. Accepts JSON:
      { "csv": "Naam, Class, Phone, Status, Next Call, Notes" }
    Or:
      { "naam": "...", "class": "...", "phone": "...",
        "status": "called", "next_call": "...", "notes": "..." }

    Existing reminder values are kept if a field is empty/missing — bot's
    auto-detected Naam/Class/Phone are pre-filled so team usually only
    needs to type Status, Next Call and Notes.
    """
    body = await request.json()
    csv_line = (body.get("csv") or "").strip()

    fields: list[str] = []
    if csv_line:
        # User-friendly comma-separated input
        fields = [f.strip() for f in csv_line.split(",")]

    def pick(idx: int, key: str) -> str | None:
        if idx < len(fields) and fields[idx]:
            return fields[idx]
        v = body.get(key)
        return v.strip() if isinstance(v, str) and v.strip() else None

    naam = pick(0, "naam")
    class_label = pick(1, "class")
    phone = pick(2, "phone")
    status = pick(3, "status") or "called"
    next_call = pick(4, "next_call")
    notes = pick(5, "notes")

    database.mark_lead_called(
        sender_id=sender_id,
        naam=naam,
        class_label=class_label,
        phone=phone,
        next_call=next_call,
        notes=notes,
        status=status,
    )

    # Push the updated lead to Google Sheet so the CRM stays in sync
    sheets_payload = database.get_lead_reminder(sender_id)
    if sheets_payload and os.getenv("GOOGLE_SHEETS_WEBHOOK"):
        try:
            import requests
            requests.post(
                os.getenv("GOOGLE_SHEETS_WEBHOOK"),
                json={
                    "phone": sheets_payload.get("phone") or "+" + sender_id,
                    "naam": sheets_payload.get("naam") or "",
                    "class": sheets_payload.get("class_label") or "",
                    "source": sheets_payload.get("source") or "",
                    "status": sheets_payload.get("status") or "called",
                    "next_call": sheets_payload.get("next_call") or "",
                    "notes": sheets_payload.get("notes") or "",
                    "is_new_lead": False,
                },
                timeout=4,
            )
        except Exception as e:
            print(f"[Sheets] Lead update sync failed: {e}")

    print(f"[Lead] +{sender_id} marked '{status}' by {user['username']} ({user['role']})")
    return {"status": "ok", "lead": sheets_payload}


@app.post("/admin/api/leads/test")
def api_create_test_lead(user: dict = Depends(verify_admin)):
    """Admin-only: creates a fake pending lead so you can test the panel + modal UI."""
    import random
    test_id = f"99{random.randint(100000000, 999999999)}"
    database.add_lead_reminder(
        sender_id=test_id,
        naam=f"Test Student {random.randint(1, 99)}",
        class_label=random.choice(["Class 9", "Class 10", "Class 11", "Junior"]),
        phone="+" + test_id,
        source=random.choice(["Meta Ad", "WhatsApp", "Referral", "Instagram"]),
    )
    print(f"[Test] Created fake pending lead +{test_id}")
    return {"status": "ok", "sender_id": test_id, "message": "Refresh /admin to see it in Pending Calls panel"}


@app.post("/admin/api/leads/{sender_id}/dismiss")
def api_lead_dismiss(sender_id: str, user: dict = Depends(verify_admin)):
    """Admin-only: dismiss a lead reminder without calling (junk/spam/etc.)."""
    database.mark_lead_called(sender_id=sender_id, status="dismissed")
    print(f"[Lead] +{sender_id} dismissed by admin {user['username']}")
    return {"status": "dismissed"}
